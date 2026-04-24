"""
DEEP ANALYSIS - Per-product accuracy backtesting
Goal: Find the OPTIMAL forecasting method for EACH product by backtesting
against known actuals, then use that method for FY26 Q2.
"""
import openpyxl
import math

wb = openpyxl.load_workbook('CFL_External Data Pack_Phase2.xlsx', data_only=True)
ws = wb['Ph.2 Data Pack-Actual Booking']
ws_big = wb['Ph.2 - Big Deal ']
ws_scms = wb['Ph.2 - SCMS']
ws_vms = wb['Ph.2 - VMS']

# =================================================================
# 1. Read all actual booking data
# =================================================================
quarters = ['FY23Q2','FY23Q3','FY23Q4','FY24Q1','FY24Q2','FY24Q3','FY24Q4','FY25Q1','FY25Q2','FY25Q3','FY25Q4','FY26Q1']

products = []
for row in range(4, 24):
    name = ws.cell(row, 2).value
    plc = ws.cell(row, 3).value
    actuals = []
    for col in range(4, 16):
        v = ws.cell(row, col).value
        actuals.append(float(v) if v is not None else 0)
    
    dp_fc = float(ws.cell(row, 17).value or 0)
    mk_fc = float(ws.cell(row, 18).value or 0)
    ds_fc = float(ws.cell(row, 19).value or 0)
    
    products.append({
        'rank': ws.cell(row, 1).value,
        'name': name,
        'plc': plc,
        'actuals': actuals,
        'dp_fc': dp_fc,
        'mk_fc': mk_fc,
        'ds_fc': ds_fc,
    })

# Read accuracy data per source per quarter
for idx, row in enumerate(range(29, 49)):
    p = products[idx]
    p['dp_acc'] = {
        'FY26Q1': float(ws.cell(row, 3).value or 0),
        'FY25Q4': float(ws.cell(row, 5).value or 0),
        'FY25Q3': float(ws.cell(row, 7).value or 0)
    }
    p['dp_bias'] = {
        'FY26Q1': float(ws.cell(row, 4).value or 0),
        'FY25Q4': float(ws.cell(row, 6).value or 0),
        'FY25Q3': float(ws.cell(row, 8).value or 0)
    }
    p['mk_acc'] = {
        'FY26Q1': float(ws.cell(row, 10).value or 0),
        'FY25Q4': float(ws.cell(row, 12).value or 0),
        'FY25Q3': float(ws.cell(row, 14).value or 0)
    }
    p['mk_bias'] = {
        'FY26Q1': float(ws.cell(row, 11).value or 0),
        'FY25Q4': float(ws.cell(row, 13).value or 0),
        'FY25Q3': float(ws.cell(row, 15).value or 0)
    }
    p['ds_acc'] = {
        'FY26Q1': float(ws.cell(row, 17).value or 0),
        'FY25Q4': float(ws.cell(row, 19).value or 0),
        'FY25Q3': float(ws.cell(row, 21).value or 0)
    }
    p['ds_bias'] = {
        'FY26Q1': float(ws.cell(row, 18).value or 0),
        'FY25Q4': float(ws.cell(row, 20).value or 0),
        'FY25Q3': float(ws.cell(row, 22).value or 0)
    }

# Read big deal data
for idx, row in enumerate(range(3, 23)):
    p = products[idx]
    p['mfg_units'] = [float(ws_big.cell(row, col).value or 0) for col in range(3, 11)]
    p['big_deals'] = [float(ws_big.cell(row, col).value or 0) for col in range(11, 19)]
    p['avg_deals'] = [float(ws_big.cell(row, col).value or 0) for col in range(19, 27)]

# Read SCMS data
scms_data = {}
for row in range(4, ws_scms.max_row + 1):
    rank = ws_scms.cell(row, 1).value
    if rank is None: continue
    channel = ws_scms.cell(row, 3).value
    vals = [float(ws_scms.cell(row, col).value or 0) for col in range(4, 17)]
    scms_data.setdefault(rank, {})[channel] = vals

# Read VMS data
vms_data = {}
for row in range(4, ws_vms.max_row + 1):
    rank = ws_vms.cell(row, 1).value
    if rank is None: continue
    vertical = ws_vms.cell(row, 3).value
    vals = [float(ws_vms.cell(row, col).value or 0) for col in range(4, 17)]
    vms_data.setdefault(rank, {})[vertical] = vals

# =================================================================
# 4. COMPREHENSIVE ANALYSIS
# =================================================================
print("=" * 120)
print("DEEP PRODUCT-BY-PRODUCT ANALYSIS")
print("=" * 120)

for p in products:
    rank = p['rank']
    act = p['actuals']
    print(f"\n{'='*120}")
    print(f"PRODUCT #{rank}: {p['name']}")
    print(f"PLC: {p['plc']}")
    print(f"{'='*120}")
    
    # Actuals timeline
    print("\nACTUALS TIMELINE:")
    for i, q in enumerate(quarters):
        print(f"  {q}: {act[i]:>12,.0f}")
    
    # Q2-specific history
    q2_vals = [act[0], act[4], act[8]]
    print(f"\nQ2 HISTORY: FY23Q2={q2_vals[0]:,.0f} | FY24Q2={q2_vals[1]:,.0f} | FY25Q2={q2_vals[2]:,.0f}")
    
    # Q2/Q1 ratio
    q2q1_ratios = []
    for q2i, q1i in [(4,3), (8,7)]:
        if act[q1i] > 0:
            q2q1_ratios.append(act[q2i] / act[q1i])
    if q2q1_ratios:
        print(f"  Q2/Q1 ratios: {[round(r,3) for r in q2q1_ratios]}")
        print(f"  Q2 estimate from FY26Q1 * avg_ratio: {round(act[11] * sum(q2q1_ratios)/len(q2q1_ratios))}")
    
    # YoY Q2 growth
    for i in range(1, len(q2_vals)):
        if q2_vals[i-1] > 0:
            g = (q2_vals[i] - q2_vals[i-1]) / q2_vals[i-1] * 100
            print(f"  Q2 YoY FY{22+i}->FY{23+i}: {g:+.1f}%")
    
    # Quarterly trend (last 4 quarters)
    last4 = act[-4:]
    print(f"\n  Last 4 qtrs: {[round(v) for v in last4]}")
    print(f"  MA4: {sum(last4)/4:,.0f}")
    
    # Expert accuracy
    print(f"\nEXPERT ACCURACY:")
    for src, key in [('Demand Planner','dp'), ('Marketing','mk'), ('Data Science','ds')]:
        acc = p[f'{key}_acc']
        bias = p[f'{key}_bias']
        w_acc = acc['FY26Q1'] * 0.5 + acc['FY25Q4'] * 0.3 + acc['FY25Q3'] * 0.2
        w_bias = bias['FY26Q1'] * 0.5 + bias['FY25Q4'] * 0.3 + bias['FY25Q3'] * 0.2
        print(f"  {src:20s}: Acc=[Q1:{acc['FY26Q1']:.1%}, Q4:{acc['FY25Q4']:.1%}, Q3:{acc['FY25Q3']:.1%}] "
              f"WtAcc={w_acc:.1%} | Bias=[Q1:{bias['FY26Q1']:+.1%}, Q4:{bias['FY25Q4']:+.1%}, Q3:{bias['FY25Q3']:+.1%}] "
              f"WtBias={w_bias:+.1%}")
        print(f"    FC for Q2: {p[f'{key}_fc']:,.0f}")
    
    # Big deal decomposition
    print(f"\nBIG DEAL DECOMP:")
    bd_q = ['FY24Q2','FY24Q3','FY24Q4','FY25Q1','FY25Q2','FY25Q3','FY25Q4','FY26Q1']
    for i, q in enumerate(bd_q):
        t = p['mfg_units'][i]
        b = p['big_deals'][i]
        a = p['avg_deals'][i]
        pct = (b/t*100) if t > 0 else 0
        print(f"  {q}: Total={t:>10,.0f} | Big={b:>10,.0f} ({pct:5.1f}%) | Avg={a:>10,.0f}")
    
    # SCMS bottom-up
    if rank in scms_data:
        print(f"\nSCMS CHANNELS:")
        s_total = 0
        for ch, vals in sorted(scms_data[rank].items()):
            # Q2 indices: 1(FY23Q2), 5(FY24Q2), 9(FY25Q2)  Q1 index: 12(FY26Q1)
            q2h = [vals[i] for i in [1,5,9] if i < len(vals)]
            fy26q1_v = vals[12] if len(vals) > 12 else 0
            s_total += fy26q1_v
            print(f"  {ch:25s}: Q2_hist={[round(v) for v in q2h]} | FY26Q1={fy26q1_v:>8,.0f}")
        print(f"  {'SCMS TOTAL':25s}: FY26Q1={s_total:>8,.0f} (Actual FY26Q1={act[11]:,.0f})")
    
    # VMS bottom-up
    if rank in vms_data:
        print(f"\nVMS VERTICALS:")
        v_total = 0
        for vt, vals in sorted(vms_data[rank].items()):
            q2h = [vals[i] for i in [1,5,9] if i < len(vals)]
            fy26q1_v = vals[12] if len(vals) > 12 else 0
            v_total += fy26q1_v
            print(f"  {vt:25s}: Q2_hist={[round(v) for v in q2h]} | FY26Q1={fy26q1_v:>8,.0f}")
        print(f"  {'VMS TOTAL':25s}: FY26Q1={v_total:>8,.0f} (Actual FY26Q1={act[11]:,.0f})")
    
    print()
