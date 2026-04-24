import openpyxl

wb = openpyxl.load_workbook('Aarya_v7_full_context_here/CFL_External Data Pack_Phase2.xlsx', data_only=True)
ws_big = wb['Ph.2 - Big Deal ']
ws = wb['Ph.2 Data Pack-Actual Booking']

print("=" * 120)
print("  AUDIT 3: BIG DEAL TREND ANALYSIS — What insights could Big Deal data give us?")
print("=" * 120)

# For each product, show big deal % of MFG across all 8 quarters
print("\nBIG DEAL CONCENTRATION (Big/MFG %) across quarters:")
print(f"{'#':<4} {'Product':<30} {'24Q2':>6} {'24Q3':>6} {'24Q4':>6} {'25Q1':>6} {'25Q2':>6} {'25Q3':>6} {'25Q4':>6} {'26Q1':>6} | {'Trend'}")
for row in range(3, 23):
    rank = ws_big.cell(row, 1).value
    name = str(ws_big.cell(row, 2).value)[:30]
    pcts = []
    for q in range(8):
        mfg = ws_big.cell(row, 3+q).value or 0
        big = ws_big.cell(row, 11+q).value or 0
        pct = big/mfg*100 if mfg > 0 else 0
        pcts.append(pct)
    # Is big deal concentration trending up or down?
    recent = pcts[4:8]   # FY25Q2 to FY26Q1
    early = pcts[0:4]    # FY24Q2 to FY25Q1
    avg_recent = sum(recent)/4 if sum(recent) > 0 else 0
    avg_early = sum(early)/4 if sum(early) > 0 else 0
    trend = "RISING" if avg_recent > avg_early + 5 else ("FALLING" if avg_recent < avg_early - 5 else "STABLE")
    
    pct_strs = [f"{p:>5.0f}%" for p in pcts]
    print(f"{rank:<4} {name:<30} {''.join(pct_strs)} | {trend}")

# KEY ANALYSIS: For products where Big Deals are a large percentage,
# the Q2 forecast should factor in whether big deals are expected to continue
print("\n\n" + "=" * 120)
print("  AUDIT 4: MFG TOTAL vs ACTUAL BOOKING — Are they the same data?")
print("=" * 120)

print("\nComparing MFG Book Units (Big Deal sheet) vs Actual Booking (main sheet):")
print(f"{'#':<4} {'Quarter':<8} {'MFG':<8} {'Actual':<8} {'Match'}")
for row in range(3, 23):
    rank = ws_big.cell(row, 1).value
    # MFG cols 3-10 = FY24Q2..FY26Q1
    # Actual cols 8-15 = FY24Q2..FY26Q1 (in main sheet, row = rank+3)
    for q_idx, (mfg_col, act_col) in enumerate([(3,8),(4,9),(5,10),(6,11),(7,12),(8,13),(9,14),(10,15)]):
        mfg_val = ws_big.cell(row, mfg_col).value or 0
        act_val = ws.cell(rank+3, act_col).value or 0
        match = "OK" if abs(mfg_val - act_val) < 2 else f"DIFF: mfg={mfg_val} act={act_val}"
        if match != "OK":
            q_names = ['24Q2','24Q3','24Q4','25Q1','25Q2','25Q3','25Q4','26Q1']
            print(f"  #{rank:<2}  {q_names[q_idx]:<8} {mfg_val:<8.0f} {act_val:<8.0f} {match}")
    # Just check #1 in detail
    if rank == 1:
        for q_idx, (mfg_col, act_col) in enumerate([(3,8),(4,9),(5,10),(6,11),(7,12),(8,13),(9,14),(10,15)]):
            mfg_val = ws_big.cell(row, mfg_col).value or 0
            act_val = ws.cell(rank+3, act_col).value or 0
            q_names = ['24Q2','24Q3','24Q4','25Q1','25Q2','25Q3','25Q4','26Q1']
            print(f"  #{rank} {q_names[q_idx]}: MFG={mfg_val:.0f} Actual={act_val:.0f}")

print("\n\n" + "=" * 120)
print("  AUDIT 5: SCMS CHANNEL SHIFT ANALYSIS — Are channels growing/declining differently?")
print("=" * 120)

ws_scms = wb['Ph.2 - SCMS']
# For product #1, show channel growth
print("\nProduct #1 SCMS Channel Q2 History (FY23Q2=C5, FY24Q2=C9, FY25Q2=C13, FY26Q1=C16):")
for r in range(4, 10):
    rank = ws_scms.cell(r, 1).value
    ch = ws_scms.cell(r, 3).value
    q2_23 = ws_scms.cell(r, 5).value or 0  # FY23Q2
    q2_24 = ws_scms.cell(r, 9).value or 0  # FY24Q2
    q2_25 = ws_scms.cell(r, 13).value or 0  # FY25Q2
    q1_26 = ws_scms.cell(r, 16).value or 0  # FY26Q1
    if rank == 1:
        growth = ((q2_25 - q2_24) / q2_24 * 100) if q2_24 > 0 else 0
        print(f"  {ch:<20} FY23Q2={q2_23:>6.0f} FY24Q2={q2_24:>6.0f} FY25Q2={q2_25:>6.0f} FY26Q1={q1_26:>6.0f} | Q2 YoY: {growth:+.1f}%")

# Check which channels have FY26Q1 data that could give Q2 signals
print("\n\nProduct #14 (SW 8P Ethernet - hockey stick growth) SCMS Channels:")
for r in range(4, ws_scms.max_row+1):
    rank = ws_scms.cell(r, 1).value
    if rank != 14:
        continue
    ch = ws_scms.cell(r, 3).value
    q2_23 = ws_scms.cell(r, 5).value or 0
    q2_24 = ws_scms.cell(r, 9).value or 0
    q2_25 = ws_scms.cell(r, 13).value or 0
    q1_26 = ws_scms.cell(r, 16).value or 0
    print(f"  {ch:<20} FY23Q2={q2_23:>6.0f} FY24Q2={q2_24:>6.0f} FY25Q2={q2_25:>6.0f} FY26Q1={q1_26:>6.0f}")

# Product Insights - what we could learn
print("\n\n" + "=" * 120)
print("  AUDIT 6: PRODUCT INSIGHTS ANALYSIS")
print("=" * 120)
ws_pi = wb['Ph.2 - Masked Product Insights ']
interesting_products = [4, 8, 9, 10]  # IP Phones
for r in range(2, ws_pi.max_row+1):
    idx = r - 1
    if idx in interesting_products:
        name = ws_pi.cell(r, 1).value
        desc = ws_pi.cell(r, 2).value
        print(f"\n  Product #{idx}: {name}")
        print(f"  Description: {desc}")

# Check accuracy formula from Glossary
print("\n\n" + "=" * 120)
print("  AUDIT 7: ACCURACY FORMULA VERIFICATION")
print("=" * 120)
ws_g = wb['Glossary']
for r in range(11, 15):
    for c in range(1, 3):
        v = ws_g.cell(r, c).value
        if v:
            print(f"  Row {r}, Col {c}: {v}")

# Check cost rank info
print("\n\nCost Rank definition:")
for r in range(14, 16):
    for c in range(1, 3):
        v = ws_g.cell(r, c).value
        if v:
            print(f"  Row {r}, Col {c}: {v}")
