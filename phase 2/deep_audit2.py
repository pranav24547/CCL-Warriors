import openpyxl

wb = openpyxl.load_workbook('Aarya_v7_full_context_here/CFL_External Data Pack_Phase2.xlsx', data_only=True)

print("=" * 100)
print("  NUMERICAL AUDIT: BIG DEAL DATA INDEX ANALYSIS")
print("=" * 100)

ws_big = wb['Ph.2 - Big Deal ']

# The code reads: range(3, 11) -> C3-C10, range(11, 19) -> C11-C18, range(19, 27) -> C19-C26
# Big deal data STARTS at row 3 (product #1), not row 4
# Let's verify what quarters map to what

print("\nBIG DEAL COLUMN MAPPING:")
print("  MFG Book Units:")
for c in range(3, 11):
    h = ws_big.cell(2, c).value
    print(f"    Col {c} (index {c-3}): {h}")

print("\n  Big Deals:")
for c in range(11, 19):
    h = ws_big.cell(2, c).value
    print(f"    Col {c} (index {c-11}): {h}")

print("\n  Avg Deals:")
for c in range(19, 27):
    h = ws_big.cell(2, c).value
    print(f"    Col {c} (index {c-19}): {h}")

# The code uses big_deals[0] and big_deals[4] for bd_fc
# big_deals[0] = C11 = 2024Q2
# big_deals[4] = C15 = 2025Q2
print("\nCODE USES:")
print("  big_deals[0] = C11 = 2024Q2")
print("  big_deals[4] = C15 = 2025Q2")
print("  avg_deals[0] = C19 = 2024Q2")
print("  avg_deals[4] = C23 = 2025Q2")

# These are the ONLY Q2 quarters being used. But there's also:
# Big Deals: C18 = 2026Q1 (current quarter!)
# Avg Deals: C26 = 2026Q1 (current quarter!)
print("\n\n  >>> UNUSED BUT VALUABLE:")
print("  big_deals[7] = C18 = 2026Q1 BIG DEALS (CURRENT QUARTER!)")
print("  avg_deals[7] = C26 = 2026Q1 AVG DEALS (CURRENT QUARTER!)")
print("  mfg_total[7] = C10 = 2026Q1 MFG TOTAL (CURRENT QUARTER!)")

# Let's see what FY26Q1 big deal data looks like
print("\n\nFY26Q1 BIG DEAL vs AVG DEAL vs MFG TOTAL:")
print(f"{'#':<4} {'Product':<40} {'MFG Q1':<8} {'Big Q1':<8} {'Avg Q1':<8} {'Big%':<8}")
for row in range(3, 23):
    rank = ws_big.cell(row, 1).value
    name = str(ws_big.cell(row, 2).value)[:40]
    mfg_q1 = ws_big.cell(row, 10).value or 0  # 2026Q1 mfg
    big_q1 = ws_big.cell(row, 18).value or 0  # 2026Q1 big deals
    avg_q1 = ws_big.cell(row, 26).value or 0  # 2026Q1 avg deals
    big_pct = f"{big_q1/mfg_q1*100:.0f}%" if mfg_q1 > 0 else "-"
    print(f"{rank:<4} {name:<40} {mfg_q1:<8} {big_q1:<8} {avg_q1:<8} {big_pct:<8}")

# MFG TOTAL: check if mfg=big+avg
print("\n\nVERIFY: MFG = Big + Avg for FY25Q2?")
for row in range(3, 23):
    rank = ws_big.cell(row, 1).value
    mfg_q2 = ws_big.cell(row, 7).value or 0   # 2025Q2 mfg
    big_q2 = ws_big.cell(row, 15).value or 0   # 2025Q2 big
    avg_q2 = ws_big.cell(row, 23).value or 0   # 2025Q2 avg
    total = big_q2 + avg_q2
    match = "OK" if abs(total - mfg_q2) < 2 else f"MISMATCH({mfg_q2} vs {total})"
    print(f"  #{rank}: mfg={mfg_q2}, big={big_q2}, avg={avg_q2}, sum={total} -> {match}")

print("\n\n" + "=" * 100)
print("  SCMS/VMS FY26Q1 SUM vs ACTUAL FY26Q1")
print("=" * 100)

ws = wb['Ph.2 Data Pack-Actual Booking']
ws_scms = wb['Ph.2 - SCMS']
ws_vms = wb['Ph.2 - VMS']

print("\nSCMS FY26Q1 channel sum vs Actual FY26Q1:")
for prod_rank in range(1, 21):
    actual_q1 = ws.cell(prod_rank + 3, 15).value or 0
    scms_sum = 0
    for r in range(4, ws_scms.max_row+1):
        if ws_scms.cell(r, 1).value == prod_rank:
            v = ws_scms.cell(r, 16).value or 0
            scms_sum += v
    diff_pct = f"{(scms_sum-actual_q1)/actual_q1*100:+.1f}%" if actual_q1 > 0 else "-"
    print(f"  #{prod_rank:<2}: Actual={actual_q1:<8.0f} SCMS_sum={scms_sum:<8.0f} diff={diff_pct}")

print("\nVMS FY26Q1 vertical sum vs Actual FY26Q1:")
for prod_rank in range(1, 21):
    actual_q1 = ws.cell(prod_rank + 3, 15).value or 0
    vms_sum = 0
    for r in range(4, ws_vms.max_row+1):
        if ws_vms.cell(r, 1).value == prod_rank:
            v = ws_vms.cell(r, 16).value or 0
            vms_sum += v
    diff_pct = f"{(vms_sum-actual_q1)/actual_q1*100:+.1f}%" if actual_q1 > 0 else "-"
    print(f"  #{prod_rank:<2}: Actual={actual_q1:<8.0f} VMS_sum={vms_sum:<8.0f} diff={diff_pct}")

# Check the accuracy section more carefully
print("\n\n" + "=" * 100)
print("  ACCURACY/BIAS SECTION: WHAT CODE READS vs ACTUAL LAYOUT")
print("=" * 100)

# Row 27 headers show: C3=FY2026Q1, C5=FY2025Q4, C7=FY2025Q3 etc.
# The code reads:
# dp: acc_q1=C3, acc_q4=C5, acc_q3=C7, bias_q1=C4, bias_q4=C6, bias_q3=C8
# mk: acc_q1=C10, acc_q4=C12, acc_q3=C14, bias_q1=C11, bias_q4=C13, bias_q3=C15
# ds: acc_q1=C17, acc_q4=C19, acc_q3=C21, bias_q1=C18, bias_q4=C20, bias_q3=C22
print("\nDP Section headers (from row 27-28):")
for c in [3,4,5,6,7,8]:
    h27 = ws.cell(27, c).value
    h28 = ws.cell(28, c).value
    print(f"  Col {c}: {h27} / {h28}")

print("\nMK Section headers:")
for c in [10,11,12,13,14,15]:
    h27 = ws.cell(27, c).value
    h28 = ws.cell(28, c).value
    print(f"  Col {c}: {h27} / {h28}")

print("\nDS Section headers:")
for c in [17,18,19,20,21,22]:
    h27 = ws.cell(27, c).value
    h28 = ws.cell(28, c).value
    print(f"  Col {c}: {h27} / {h28}")

# Check col 9 and 16 in accuracy section - are there gaps?
print("\nGap columns in accuracy section:")
for c in [9, 16]:
    h26 = ws.cell(26, c).value
    h27 = ws.cell(27, c).value
    h28 = ws.cell(28, c).value
    v29 = ws.cell(29, c).value
    print(f"  Col {c}: row26={h26}, row27={h27}, row28={h28}, row29={v29}")
