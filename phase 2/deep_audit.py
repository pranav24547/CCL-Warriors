import openpyxl

wb = openpyxl.load_workbook('Aarya_v7_full_context_here/CFL_External Data Pack_Phase2.xlsx', data_only=True)

print("=" * 120)
print("  DEEP AUDIT: ALL DATA IN PHASE 2 DATAPACK")
print("=" * 120)

# ============================================================
# SHEET 1: Actual Booking - Full column map
# ============================================================
ws = wb['Ph.2 Data Pack-Actual Booking']
print("\n### SHEET: Ph.2 Data Pack-Actual Booking ###")
print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")

# Print ALL row headers to understand the full layout
print("\n  SECTION 1: HEADER ROWS (rows 1-3)")
for r in range(1, 4):
    for c in range(1, ws.max_column+1):
        v = ws.cell(r, c).value
        if v is not None:
            print(f"    Row {r}, Col {c}: {v}")

print("\n  SECTION 2: PRODUCT DATA (rows 4-23)")
for r in range(4, 6):  # Just first 2 products
    print(f"  Product row {r}:")
    for c in range(1, ws.max_column+1):
        v = ws.cell(r, c).value
        if v is not None:
            print(f"    Col {c}: {v}")

# Check what's in rows 24-48 (accuracy/bias section)
print("\n  SECTION 3: ROWS 24-28 (gap/headers)")
for r in range(24, 30):
    vals = []
    for c in range(1, ws.max_column+1):
        v = ws.cell(r, c).value
        if v is not None:
            vals.append(f"C{c}={v}")
    if vals:
        print(f"    Row {r}: {vals[:15]}")

print("\n  SECTION 4: ACCURACY/BIAS DATA (rows 29-48)")
for r in range(29, 31):  # First 2 products
    print(f"  Acc/Bias row {r}:")
    for c in range(1, ws.max_column+1):
        v = ws.cell(r, c).value
        if v is not None:
            print(f"    Col {c}: {v}")

# ============================================================
# SHEET 2: Big Deal - All columns
# ============================================================
ws_big = wb['Ph.2 - Big Deal ']
print("\n\n### SHEET: Ph.2 - Big Deal ###")
print(f"  Rows: {ws_big.max_row}, Cols: {ws_big.max_column}")

print("\n  HEADERS:")
for r in range(1, 3):
    for c in range(1, ws_big.max_column+1):
        v = ws_big.cell(r, c).value
        if v is not None:
            print(f"    Row {r}, Col {c}: {v}")

print("\n  FIRST PRODUCT DATA (row 3):")
for c in range(1, ws_big.max_column+1):
    v = ws_big.cell(3, c).value
    if v is not None:
        print(f"    Col {c}: {v}")

# Check columns 18-26 (Avg Deals has more data?)
print("\n  COLUMN RANGES:")
print(f"    MFG Book Units: C3-C10 (8 quarters)")
print(f"    Big Deals: C11-C18 (8 quarters)")
print(f"    Avg Deals: C19-C26 (8 quarters)")

# Verify by checking row 2 header labels
for c in range(17, ws_big.max_column+1):
    v = ws_big.cell(2, c).value
    if v is not None:
        print(f"    Row 2, Col {c}: {v}")

# Print all data for product #1 to see what we have
print("\n  ALL Big Deal data for Product #1 (row 3):")
for c in range(3, ws_big.max_column+1):
    v = ws_big.cell(3, c).value
    print(f"    Col {c}: {v}")

# ============================================================
# SHEET 3: SCMS - Structure audit
# ============================================================
ws_scms = wb['Ph.2 - SCMS']
print("\n\n### SHEET: Ph.2 - SCMS ###")
print(f"  Rows: {ws_scms.max_row}, Cols: {ws_scms.max_column}")

# Count unique channels per product
scms_channels = {}
for r in range(4, ws_scms.max_row+1):
    rank = ws_scms.cell(r, 1).value
    ch = ws_scms.cell(r, 3).value
    if rank is not None:
        scms_channels.setdefault(rank, []).append(ch)

print("\n  SCMS Channels per Product:")
for rank, channels in sorted(scms_channels.items()):
    print(f"    Product #{rank}: {len(channels)} channels: {channels}")

# Check if SCMS has FY26Q1 data (col 16)
print("\n  SCMS FY26Q1 data sample (col 16):")
for r in range(4, 10):
    rank = ws_scms.cell(r, 1).value
    ch = ws_scms.cell(r, 3).value
    q1_26 = ws_scms.cell(r, 16).value
    print(f"    Row {r}: Product #{rank}, Channel={ch}, FY26Q1={q1_26}")

# ============================================================
# SHEET 4: VMS - Structure audit
# ============================================================
ws_vms = wb['Ph.2 - VMS']
print("\n\n### SHEET: Ph.2 - VMS ###")
print(f"  Rows: {ws_vms.max_row}, Cols: {ws_vms.max_column}")

# Count unique verticals per product
vms_verticals = {}
for r in range(4, ws_vms.max_row+1):
    rank = ws_vms.cell(r, 1).value
    vt = ws_vms.cell(r, 3).value
    if rank is not None:
        vms_verticals.setdefault(rank, []).append(vt)

print("\n  VMS Verticals per Product:")
for rank, verticals in sorted(vms_verticals.items()):
    print(f"    Product #{rank}: {len(verticals)} verticals")

# ============================================================
# SHEET 5: Product Insights
# ============================================================
ws_pi = wb['Ph.2 - Masked Product Insights ']
print("\n\n### SHEET: Ph.2 - Masked Product Insights ###")
print(f"  Rows: {ws_pi.max_row}, Cols: {ws_pi.max_column}")
for r in range(2, ws_pi.max_row+1):
    name = ws_pi.cell(r, 1).value
    desc = ws_pi.cell(r, 2).value
    if name:
        print(f"  #{r-1}: {name[:50]}... -> {desc[:80] if desc else 'NO DESC'}...")

# ============================================================
# SHEET 6: Glossary
# ============================================================
ws_g = wb['Glossary']
print("\n\n### SHEET: Glossary ###")
print(f"  Rows: {ws_g.max_row}, Cols: {ws_g.max_column}")
for r in range(1, ws_g.max_row+1):
    vals = []
    for c in range(1, ws_g.max_column+1):
        v = ws_g.cell(r, c).value
        if v is not None:
            vals.append(f"C{c}={v}")
    if vals:
        print(f"    Row {r}: {vals}")

# ============================================================
# CROSS-REFERENCE: What the CODE reads
# ============================================================
print("\n\n" + "=" * 120)
print("  CROSS-REFERENCE: WHAT V7.0 CODE ACTUALLY READS")
print("=" * 120)

print("""
  FROM 'Ph.2 Data Pack-Actual Booking':
    - Rows 4-23: rank(C1), name(C2), plc(C3), actuals(C4-C15), dp_fc(C17), mk_fc(C18), ds_fc(C19)
    - Rows 29-48: dp_acc(C3,5,7), dp_bias(C4,6,8), mk_acc(C10,12,14), mk_bias(C11,13,15), ds_acc(C17,19,21), ds_bias(C18,20,22)
    >>> COLUMN 16 (Your Forecast FY26 Q2) = OUTPUT COLUMN, not read
    >>> COLUMN 20-22 in rows 4-23 = NOT READ - what's there?
  
  FROM 'Ph.2 - Big Deal ':
    - Rows 3-22: mfg_total(C3-C10), big_deals(C11-C18), avg_deals(C19-C26)
    >>> ONLY uses big_deals[0]=C11 & big_deals[4]=C15, avg_deals[0]=C19 & avg_deals[4]=C23
    >>> That means it only uses FY24Q2 and FY25Q2 from big/avg deals
    >>> IGNORING: C12-14 (Q3,Q4,Q1), C16-18 (Q3,Q4,Q1), C20-22 (Q3,Q4,Q1), C24-26 (Q3,Q4,Q1)
    >>> And mfg_total is read but NEVER USED in calculations
  
  FROM 'Ph.2 - SCMS':
    - Used in bottom_up_q2() function
    - Reads cols 4-16 (13 quarters of data)
    - Uses Q2 indices [1, 5, 9] = FY23Q2, FY24Q2, FY25Q2
    >>> But ONLY as validation signal, NOT in structural median
    >>> FY26Q1 (col 16) data EXISTS but is NOT used for ratio-based SCMS forecasting
  
  FROM 'Ph.2 - VMS':
    - Same as SCMS: used in bottom_up_q2(), only Q2 history
    >>> ONLY as validation signal, NOT in structural median
    >>> FY26Q1 (col 16) data EXISTS but is NOT used
  
  FROM 'Ph.2 - Masked Product Insights ':
    >>> COMPLETELY UNUSED - product descriptions not read at all
  
  FROM 'Glossary':
    >>> Reference only - accuracy formula definition
""")

# Now let's check what's actually in cols 20-22 of the main sheet rows 4-23
print("  CHECKING: Cols 20-22 in main sheet rows 4-23:")
for r in range(3, 6):
    for c in range(16, ws.max_column+1):
        v = ws.cell(r, c).value
        if v is not None:
            print(f"    Row {r}, Col {c}: {v}")

# Check Big Deal header row 2 more carefully
print("\n  BIG DEAL FULL HEADER (Row 2):")
for c in range(1, ws_big.max_column+1):
    v = ws_big.cell(2, c).value
    if v is not None:
        print(f"    Col {c}: {v}")

# Check if there's more data after row 22 in Big Deal
print("\n  BIG DEAL ROWS AFTER 22:")
for r in range(22, ws_big.max_row+1):
    vals = []
    for c in range(1, 5):
        v = ws_big.cell(r, c).value
        if v is not None:
            vals.append(f"C{c}={v}")
    if vals:
        print(f"    Row {r}: {vals}")
