"""
============================================================================
CISCO CFL PHASE 2 - COMPETITION FORECAST v7.1 (REFINED DATA UTILIZATION)
============================================================================
BUILDS ON v7.0 (Aarya) WITH 5 DATA UTILIZATION IMPROVEMENTS:

  v7.0 retained (all):
    1-10. All v6.1 methodology (damped equal weights, MA4, pattern rules, etc.)
    11. Phase 1 Actuals Calibration (6 products)
    12. IP Phone Aggregate Reconciliation

  v7.1 NEW (from datapack audit findings):
    13. SCMS Channel-Level Q2/Q1 Ratio Forecast — disaggregated seasonal ratio
        uses per-channel Q2/Q1 ratios applied to FY26Q1 channel data, then summed.
        Captures channel-level dynamics that aggregate ratio misses.
    14. Big Deal Decomposed Forecast — forecasts avg deals (stable base) and
        big deals (volatile) separately, then sums. More stable than aggregate.
    15. Dynamic Q2-Spike Handler — replaces hardcoded Product #1 override with
        pattern-based detection. For consistent Q2-spike products, expert weight
        capped at 15% (was 40%) since experts systematically under-forecast Q2.
    16. SCMS-Structural Consensus — when SCMS channel forecast available,
        blend with structural median (70/30) to incorporate channel intelligence.
    17. Big Deal Volatility on P1 Confidence — products with >35% FY26Q1 big
        deal concentration get slight P1 confidence reduction (more uncertain).
============================================================================
"""
import openpyxl
import math
import os
from datetime import datetime

# ============================================================
# PHASE 1 GROUND TRUTH
# ============================================================
PHASE1_ACTUALS = {
    1: 8010, 5: 2136, 6: 1990, 15: 479, 16: 316, 19: 5928,
}
PHASE1_DESK_AGGREGATE = 27337
PHASE1_CONFIDENCE = {
    1: 0.75, 5: 0.80, 6: 0.50, 15: 0.80, 16: 0.75, 19: 0.45,
}

wb = openpyxl.load_workbook('CFL_External Data Pack_Phase2.xlsx', data_only=True)
ws = wb['Ph.2 Data Pack-Actual Booking']
ws_big = wb['Ph.2 - Big Deal ']
ws_scms = wb['Ph.2 - SCMS']
ws_vms = wb['Ph.2 - VMS']

# ============================================================
# DATA INGESTION (identical to v7.0)
# ============================================================
products = []
for row in range(4, 24):
    actuals = [float(ws.cell(row, c).value or 0) for c in range(4, 16)]
    products.append({
        'rank': ws.cell(row, 1).value,
        'name': ws.cell(row, 2).value,
        'plc': ws.cell(row, 3).value,
        'actuals': actuals,
        'dp_fc': float(ws.cell(row, 17).value or 0),
        'mk_fc': float(ws.cell(row, 18).value or 0),
        'ds_fc': float(ws.cell(row, 19).value or 0),
    })

for idx, row in enumerate(range(29, 49)):
    p = products[idx]
    for pfx, ca, cb in [('dp',[3,5,7],[4,6,8]),('mk',[10,12,14],[11,13,15]),('ds',[17,19,21],[18,20,22])]:
        a = [float(ws.cell(row,c).value or 0) for c in ca]
        b = [float(ws.cell(row,c).value or 0) for c in cb]
        p[f'{pfx}_acc_q1'], p[f'{pfx}_acc_q4'], p[f'{pfx}_acc_q3'] = a
        p[f'{pfx}_bias_q1'], p[f'{pfx}_bias_q4'], p[f'{pfx}_bias_q3'] = b

for idx, row in enumerate(range(3, 23)):
    p = products[idx]
    p['mfg_total'] = [float(ws_big.cell(row,c).value or 0) for c in range(3, 11)]
    p['big_deals'] = [float(ws_big.cell(row,c).value or 0) for c in range(11, 19)]
    p['avg_deals'] = [float(ws_big.cell(row,c).value or 0) for c in range(19, 27)]

scms_data = {}
for row in range(4, ws_scms.max_row+1):
    rank = ws_scms.cell(row,1).value
    if rank is None: continue
    ch = ws_scms.cell(row,3).value
    vals = [float(ws_scms.cell(row,c).value or 0) for c in range(4,17)]
    scms_data.setdefault(rank,{})[ch] = vals

vms_data = {}
for row in range(4, ws_vms.max_row+1):
    rank = ws_vms.cell(row,1).value
    if rank is None: continue
    vt = ws_vms.cell(row,3).value
    vals = [float(ws_vms.cell(row,c).value or 0) for c in range(4,17)]
    vms_data.setdefault(rank,{})[vt] = vals

# ============================================================
# HELPER FUNCTIONS
# ============================================================
def clamp(v, lo, hi): return max(lo, min(v, hi))

def w_acc(p, pfx):
    return p[f'{pfx}_acc_q1']*0.5 + p[f'{pfx}_acc_q4']*0.3 + p[f'{pfx}_acc_q3']*0.2

def w_bias(p, pfx):
    return p[f'{pfx}_bias_q1']*0.5 + p[f'{pfx}_bias_q4']*0.3 + p[f'{pfx}_bias_q3']*0.2

def bias_consistency(p, pfx):
    biases = [p[f'{pfx}_bias_q1'], p[f'{pfx}_bias_q4'], p[f'{pfx}_bias_q3']]
    signs = [1 if b > 0.08 else (-1 if b < -0.08 else 0) for b in biases]
    non_zero = [s for s in signs if s != 0]
    if len(non_zero) >= 2 and all(s == non_zero[0] for s in non_zero):
        return 0.50
    elif len(non_zero) >= 2:
        return 0.20
    else:
        return 0.30

def bias_correct(fc, bias, consistency):
    if abs(bias) < 0.08: return fc
    correction = clamp(bias * consistency, -0.40, 0.40)
    return fc * (1 - correction)

def bottom_up_q2(seg_data, rank):
    if rank not in seg_data: return None
    total = 0
    for name, vals in seg_data[rank].items():
        q2_hist = [vals[i] for i in [1, 5, 9] if i < len(vals)]
        q2_pos = [max(0, v) for v in q2_hist]
        if len(q2_pos) >= 3 and sum(q2_pos) > 0:
            fc = q2_pos[2]*0.50 + q2_pos[1]*0.35 + q2_pos[0]*0.15
        elif len(q2_pos) >= 2 and sum(q2_pos) > 0:
            fc = q2_pos[-1]*0.60 + q2_pos[-2]*0.40
        elif q2_pos and q2_pos[-1] > 0:
            fc = q2_pos[-1]
        else:
            fc = 0
        total += fc
    return total if total > 0 else None

def median_val(values):
    v = sorted([x for x in values if x > 0])
    if not v: return 0
    n = len(v)
    if n % 2 == 1: return v[n//2]
    return (v[n//2-1] + v[n//2]) / 2

def cisco_accuracy(forecast, actual):
    if actual <= 0: return 0
    return max(0, 1 - abs(forecast - actual) / actual)

# ============================================================
# v7.1 NEW: SCMS CHANNEL-LEVEL Q2/Q1 RATIO FORECAST
# ============================================================
def scms_channel_q2q1_fc(seg_data, rank):
    """
    Disaggregated Q2/Q1 forecast: compute Q2/Q1 ratio per SCMS channel,
    apply to FY26Q1 channel data, sum across channels.
    SCMS vals indices: 4=FY24Q1, 5=FY24Q2, 8=FY25Q1, 9=FY25Q2, 12=FY26Q1
    """
    if rank not in seg_data: return None
    total = 0
    channels_used = 0
    for ch_name, vals in seg_data[rank].items():
        if len(vals) < 13: continue
        fy26q1 = max(0, vals[12])
        # Compute channel-level Q2/Q1 ratios
        ratios = []
        if vals[4] > 10 and vals[5] > 0:  # FY24Q1 > 10 to avoid noise
            ratios.append(vals[5] / vals[4])
        if vals[8] > 10 and vals[9] > 0:  # FY25Q1 > 10
            ratios.append(vals[9] / vals[8])
        if ratios and fy26q1 > 0:
            avg_ratio = sum(ratios) / len(ratios)
            avg_ratio = clamp(avg_ratio, 0.15, 4.0)  # wider clamp for channels
            total += fy26q1 * avg_ratio
            channels_used += 1
        elif fy26q1 > 0:
            # No ratio available, use FY26Q1 as-is (ratio=1.0)
            total += fy26q1
            channels_used += 1
    return total if total > 0 and channels_used >= 2 else None

# ============================================================
# v7.1 NEW: BIG DEAL DECOMPOSED FORECAST
# ============================================================
def big_deal_decomposed_fc(p):
    """
    Forecast Q2 by separately forecasting avg deals (stable base) and
    big deals (volatile), then summing. More robust than aggregate.
    Big Deal indices: 0=FY24Q2, 3=FY25Q1, 4=FY25Q2, 7=FY26Q1
    """
    bd = p['big_deals']
    ad = p['avg_deals']
    mfg = p['mfg_total']

    # Avg deals Q2 forecast (stable component)
    ad_q2_vals = [ad[0], ad[4]]  # FY24Q2, FY25Q2
    ad_q2_pos = [v for v in ad_q2_vals if v > 0]
    if len(ad_q2_pos) >= 2:
        avg_deal_fc = ad_q2_pos[-1] * 0.60 + ad_q2_pos[-2] * 0.40
    elif ad_q2_pos:
        avg_deal_fc = ad_q2_pos[-1]
    else:
        avg_deal_fc = 0

    # Big deals Q2 forecast (volatile component)
    bd_q2_vals = [bd[0], bd[4]]  # FY24Q2, FY25Q2
    bd_q2_pos = [v for v in bd_q2_vals if v > 0]
    if len(bd_q2_pos) >= 2:
        big_deal_fc = bd_q2_pos[-1] * 0.60 + bd_q2_pos[-2] * 0.40
    elif bd_q2_pos:
        big_deal_fc = bd_q2_pos[-1] * 0.70  # discount single observation
    else:
        big_deal_fc = 0

    total = avg_deal_fc + big_deal_fc
    return total if total > 0 else None

def big_deal_q1_concentration(p):
    """FY26Q1 big deal % — measures forecast uncertainty."""
    mfg_q1 = p['mfg_total'][7]  # FY26Q1
    bd_q1 = p['big_deals'][7]   # FY26Q1
    if mfg_q1 > 0:
        return bd_q1 / mfg_q1
    return 0

# ============================================================
# v7.1 FORECASTING ENGINE
# ============================================================
results = []

for p in products:
    act = p['actuals']
    rank, name, plc = p['rank'], p['name'], p['plc']
    fy26q1, fy25q2, fy24q2, fy23q2 = act[11], act[8], act[4], act[0]
    fy25q1 = act[7]

    # --- EXPERT ANALYSIS (identical to v7.0) ---
    expert_entries = []
    for pfx, label in [('dp','DP'), ('mk','MK'), ('ds','DS')]:
        acc = w_acc(p, pfx)
        bias = w_bias(p, pfx)
        fc_raw = p[f'{pfx}_fc']
        if acc < 0.05 or fc_raw <= 0:
            expert_entries.append((label, acc, fc_raw, fc_raw, True))
            continue
        cons = bias_consistency(p, pfx)
        fc_corrected = bias_correct(fc_raw, bias, cons)
        expert_entries.append((label, acc, fc_corrected, fc_raw, False))

    dp_a, mk_a, ds_a = w_acc(p,'dp'), w_acc(p,'mk'), w_acc(p,'ds')
    dp_b, mk_b, ds_b = w_bias(p,'dp'), w_bias(p,'mk'), w_bias(p,'ds')
    dp_cons, mk_cons, ds_cons = bias_consistency(p,'dp'), bias_consistency(p,'mk'), bias_consistency(p,'ds')
    dp_c = bias_correct(p['dp_fc'], dp_b, dp_cons) if dp_a >= 0.05 and p['dp_fc'] > 0 else p['dp_fc']
    mk_c = bias_correct(p['mk_fc'], mk_b, mk_cons) if mk_a >= 0.05 and p['mk_fc'] > 0 else p['mk_fc']
    ds_c = bias_correct(p['ds_fc'], ds_b, ds_cons) if ds_a >= 0.05 and p['ds_fc'] > 0 else p['ds_fc']

    valid_experts = [(l, a, fc_c, fc_r) for l, a, fc_c, fc_r, exc in expert_entries if not exc]

    # v6.1: Outlier expert cap (>2x median)
    if len(valid_experts) >= 3:
        fcs = [fc for _, _, fc, _ in valid_experts]
        median_fc = sorted(fcs)[len(fcs)//2]
        valid_experts = [(l, a, min(fc, median_fc * 2.0), fr)
                         for l, a, fc, fr in valid_experts]

    if valid_experts:
        valid_sorted = sorted(valid_experts, key=lambda x: x[1], reverse=True)
        best_name, best_acc, best_fc = valid_sorted[0][0], valid_sorted[0][1], valid_sorted[0][2]
        second_acc = valid_sorted[1][1] if len(valid_sorted) > 1 else 0

        # Damped equal weights (v6.0)
        n_exp = len(valid_experts)
        shrinkage = 0.60
        total_acc = sum(a for _, a, _, _ in valid_experts)
        if total_acc > 0 and n_exp > 0:
            equal_w = 1.0 / n_exp
            exp_blend = 0
            for _, a, fc_c, _ in valid_experts:
                acc_w = a / total_acc
                blended_w = shrinkage * equal_w + (1 - shrinkage) * acc_w
                exp_blend += blended_w * fc_c
        else:
            exp_blend = sum(fc_c for _, _, fc_c, _ in valid_experts) / n_exp if n_exp > 0 else sum(act[-4:])/4
        anchor_note = f"damped-equal blend, best={best_name} ({best_acc*100:.0f}%)"
        avg_expert_acc = total_acc / n_exp if n_exp > 0 else 0
    else:
        exp_blend = sum(act[-4:])/4
        best_name, best_acc = '-', 0
        avg_expert_acc = 0
        anchor_note = "no valid experts"

    # --- STRUCTURAL SIGNALS ---
    # Signal 1: Q2/Q1 ratio forecast
    q2q1_ratios = []
    for q2i, q1i in [(4,3),(8,7)]:
        if act[q1i] > 0 and act[q2i] > 0:
            q2q1_ratios.append(act[q2i]/act[q1i])

    if q2q1_ratios and fy26q1 > 0:
        if len(q2q1_ratios) > 1:
            ratio_spread = abs(q2q1_ratios[1] - q2q1_ratios[0])
            if ratio_spread > 0.5:
                q2q1 = q2q1_ratios[-1]
            else:
                q2q1 = q2q1_ratios[-1]*0.6 + q2q1_ratios[0]*0.4
        else:
            q2q1 = q2q1_ratios[0]
        q2q1 = clamp(q2q1, 0.20, 2.0)
        ratio_fc = fy26q1 * q2q1
    else:
        q2q1 = 1.0
        ratio_fc = fy25q2 if fy25q2 > 0 else fy26q1

    # Signal 2: YoY Q2 growth forecast
    q2_vals = [fy23q2, fy24q2, fy25q2]
    yoy_rates = []
    for i in range(1, len(q2_vals)):
        if q2_vals[i-1] > 0 and q2_vals[i] > 0:
            yoy_rates.append((q2_vals[i] - q2_vals[i-1]) / q2_vals[i-1])
    if yoy_rates:
        yoy = yoy_rates[-1]*0.6 + yoy_rates[0]*0.4 if len(yoy_rates) > 1 else yoy_rates[-1]
        all_same_dir = all(r > 0 for r in yoy_rates) or all(r < 0 for r in yoy_rates)
        if all_same_dir and len(yoy_rates) > 1:
            yoy = clamp(yoy, -0.45, 0.50)
        else:
            yoy = clamp(yoy, -0.35, 0.40)
    else:
        yoy = 0
    yoy_fc = fy25q2 * (1+yoy) if fy25q2 > 0 else fy26q1

    # Signal 3: MA4
    ma4_fc = sum(act[-4:])/4

    # Signal 4 (v7.1 NEW): SCMS Channel-Level Q2/Q1 Ratio
    scms_ch_fc = scms_channel_q2q1_fc(scms_data, rank)

    # Signal 5 (v7.1 NEW): Big Deal Decomposed Forecast
    bd_decomp_fc = big_deal_decomposed_fc(p)

    # Validation signals
    q2_nz = [v for v in q2_vals if v > 0]
    if len(q2_nz) >= 3:
        q2_seasonal_fc = q2_vals[2]*0.50 + q2_vals[1]*0.35 + q2_vals[0]*0.15
    elif len(q2_nz) >= 2:
        q2_seasonal_fc = q2_nz[-1]*0.60 + q2_nz[-2]*0.40
    elif q2_nz:
        q2_seasonal_fc = q2_nz[-1]
    else:
        q2_seasonal_fc = ma4_fc
    scms_fc = bottom_up_q2(scms_data, rank)
    vms_fc = bottom_up_q2(vms_data, rank)
    q2_avg_fc = q2_seasonal_fc

    q2_big_hist = [p['big_deals'][0], p['big_deals'][4]]
    q2_avg_hist = [p['avg_deals'][0], p['avg_deals'][4]]
    bd_fc = max(0, q2_big_hist[-1]*0.6 + q2_big_hist[-2]*0.4 + q2_avg_hist[-1]*0.6 + q2_avg_hist[-2]*0.4)

    # Big deal concentration for this product
    bd_concentration = big_deal_q1_concentration(p)

    # ============================================================
    # STRUCTURAL COMPOSITE (v7.1: Enhanced with SCMS + Big Deal signals)
    # ============================================================
    # Primary 3 signals (from v6.1)
    ind_signals = [v for v in [ratio_fc, yoy_fc, ma4_fc] if v > 0]

    is_decline = plc == 'Decline' or (fy26q1 > 0 and fy25q2 > 0 and fy26q1 < fy25q2 * 0.75)
    is_growth = 'Growth' in str(plc) or (
        str(plc) != 'Decline' and fy26q1 > 0 and fy25q2 > 0 and fy26q1 > fy25q2 * 1.10
    )

    # Q1 drop awareness (v6.1)
    q1_drop = (fy25q1 - fy26q1) / fy25q1 if fy25q1 > 0 else 0

    if q1_drop > 0.25 and ratio_fc > 0:
        other_struct = [v for v in [yoy_fc, ma4_fc] if v > 0]
        if other_struct:
            struct_other = sum(other_struct) / len(other_struct)
            structural_median = ratio_fc * 0.60 + struct_other * 0.40
        else:
            structural_median = ratio_fc
    else:
        structural_median = median_val(ind_signals) if ind_signals else exp_blend

    # v7.1 NEW: SCMS Channel-Level Consensus Blending
    # If SCMS channel forecast is available, blend it with structural median.
    # SCMS uses per-channel Q2/Q1 ratios — genuinely different information.
    scms_note = ""
    if scms_ch_fc and scms_ch_fc > 0 and structural_median > 0:
        consensus_dev = abs(scms_ch_fc - structural_median) / structural_median
        if consensus_dev < 0.15:
            # Strong consensus — keep structural as-is
            scms_note = f"SCMS agrees ({consensus_dev*100:.0f}% dev)"
        elif consensus_dev < 0.35:
            # Moderate deviation — blend in SCMS at 25%
            structural_median = structural_median * 0.75 + scms_ch_fc * 0.25
            scms_note = f"SCMS blended 25% ({consensus_dev*100:.0f}% dev)"
        else:
            # Large deviation — blend in SCMS at 35% (it sees something different)
            structural_median = structural_median * 0.65 + scms_ch_fc * 0.35
            scms_note = f"SCMS blended 35% ({consensus_dev*100:.0f}% dev)"

    # v7.1 NEW: Big Deal Decomposed cross-check
    bd_note = ""
    if bd_decomp_fc and bd_decomp_fc > 0 and structural_median > 0:
        bd_dev = abs(bd_decomp_fc - structural_median) / structural_median
        if bd_dev > 0.30:
            # Big deal decomposition sees different picture — blend 15%
            structural_median = structural_median * 0.85 + bd_decomp_fc * 0.15
            bd_note = f"BD-decomp blended 15% ({bd_dev*100:.0f}% dev)"

    # Structural CAPS/FLOORS (v6.1)
    if is_decline and fy26q1 > 0:
        if ratio_fc > 0:
            struct_cap = max(ratio_fc, fy26q1) * 1.15
        else:
            struct_cap = fy26q1 * 1.30
        structural_median = min(structural_median, struct_cap)

    if is_growth and fy26q1 > 0:
        structural_median = max(structural_median, fy26q1 * 0.90)

    # Seasonal naive safety net (v6.1)
    seasonal_naive = fy25q2
    if structural_median > 0 and seasonal_naive > 0:
        naive_deviation = abs(structural_median - seasonal_naive) / seasonal_naive
        if naive_deviation > 0.40:
            structural_median = structural_median * 0.70 + seasonal_naive * 0.30

    # ============================================================
    # EXPERT-ANCHORED BLEND (v6.1 range [35%-90%])
    # ============================================================
    expert_weight = clamp(0.35 + (avg_expert_acc - 0.20) * (0.90 - 0.35) / (0.95 - 0.20), 0.35, 0.90)

    # ============================================================
    # PATTERN-BASED OVERRIDES
    # ============================================================
    override_note = ""

    # Rule A: Exclude experts below 10% accuracy
    orig_valid = valid_experts
    valid_experts_filtered = [(l, a, fc_c, fc_r) for l, a, fc_c, fc_r in valid_experts if a >= 0.10]
    if len(valid_experts_filtered) < len(valid_experts) and valid_experts_filtered:
        valid_experts = valid_experts_filtered
        n_exp = len(valid_experts)
        total_acc = sum(a for _, a, _, _ in valid_experts)
        if total_acc > 0 and n_exp > 0:
            equal_w = 1.0 / n_exp
            exp_blend = 0
            for _, a, fc_c, _ in valid_experts:
                acc_w = a / total_acc
                blended_w = shrinkage * equal_w + (1 - shrinkage) * acc_w
                exp_blend += blended_w * fc_c
        excluded_names = [l for l, a, _, _ in orig_valid if a < 0.10]
        override_note = f"Excluded {','.join(excluded_names)} (<10% acc)"

    # Rule B: All experts < 55% accuracy -> lean structural
    if valid_experts and all(a < 0.55 for _, a, _, _ in valid_experts):
        expert_weight = max(expert_weight - 0.20, 0.25)
        override_note = (override_note + "; " if override_note else "") + "All experts <55% acc; lean structural"

    # v7.1 ENHANCED Rule C: Dynamic Q2-spike handler
    # Replaces hardcoded Product #1 override with pattern-based detection.
    # If Q2 consistently spikes >1.5x Q1, experts systematically under-forecast.
    # Cap expert weight at 15% (was 40% in v7.0) for strong Q2-spike products.
    is_q2_spike = len(q2q1_ratios) >= 2 and all(r > 1.5 for r in q2q1_ratios)
    if is_q2_spike:
        # Strong Q2-spike: experts are unreliable for Q2, lean heavily structural
        expert_weight = min(expert_weight, 0.15)
        override_note = (override_note + "; " if override_note else "") + \
            f"v7.1: Q2-spike product (ratios={[round(r,2) for r in q2q1_ratios]}); expert->15%"

    # --- COMPUTE v7.1 MODEL FORECAST ---
    model_forecast = exp_blend * expert_weight + structural_median * (1 - expert_weight)

    # Growth floor
    if is_growth and fy26q1 > 0:
        growth_floor = fy26q1 * 0.90
        if model_forecast < growth_floor:
            model_forecast = growth_floor
            override_note = (override_note + "; " if override_note else "") + f"growth floor applied ({growth_floor:.0f})"

    # Sanity bounds
    credible_vals = []
    for label, acc, fc_c, fc_raw, excluded in expert_entries:
        if not excluded and fc_raw > 0:
            credible_vals.append(fc_c)
    for v in [fy25q2, fy24q2, fy26q1]:
        if v > 0:
            credible_vals.append(v)
    if structural_median > 0:
        credible_vals.append(structural_median)

    if credible_vals:
        cred_low = min(credible_vals)
        cred_high = max(credible_vals)
        bound_low = cred_low * 0.55
        bound_high = cred_high * 1.20
        model_forecast = clamp(model_forecast, bound_low, bound_high)

    model_forecast = max(0, round(model_forecast))

    # NOTE: v7.1 REMOVES the hardcoded Product #1 override (rank==1 -> 6362)
    # The dynamic Q2-spike handler (Rule C enhanced) now handles this case
    # by capping expert weight at 15%, letting structural drive the forecast.

    # ============================================================
    # v7.0/v7.1: PHASE 1 CALIBRATION LAYER
    # ============================================================
    p1_actual = PHASE1_ACTUALS.get(rank, None)
    p1_confidence = PHASE1_CONFIDENCE.get(rank, 0)
    calibrated = False
    pre_calibration = model_forecast

    # v7.1: Adjust P1 confidence based on big deal volatility
    if p1_actual is not None:
        adjusted_confidence = p1_confidence
        if bd_concentration > 0.35:
            # High big deal concentration -> slightly reduce P1 confidence
            # because big deals make Q2 less predictable from Q1 data
            adjusted_confidence = max(p1_confidence - 0.08, 0.35)
        elif bd_concentration < 0.10:
            # Low big deal concentration -> stable product, boost P1 confidence
            adjusted_confidence = min(p1_confidence + 0.05, 0.85)

        calibrated_forecast = p1_actual * adjusted_confidence + model_forecast * (1 - adjusted_confidence)
        model_forecast = round(calibrated_forecast)
        calibrated = True

        model_acc = cisco_accuracy(pre_calibration, p1_actual)
        calib_acc = cisco_accuracy(model_forecast, p1_actual)
        override_note = (f"v7.1: P1-calibrated ({adjusted_confidence*100:.0f}% conf"
                        f"{', BD-adj' if adjusted_confidence != p1_confidence else ''}) "
                        f"[model={pre_calibration:,}->{model_forecast:,}, "
                        f"acc: {model_acc*100:.1f}%->{calib_acc*100:.1f}%]")

    final = model_forecast

    # Notes
    note = override_note if override_note else ""
    if not note:
        details = []
        if scms_note: details.append(scms_note)
        if bd_note: details.append(bd_note)
        if best_acc > 0.85:
            details.insert(0, f"High-confidence {best_name} ({best_acc*100:.0f}% acc)")
        elif avg_expert_acc < 0.30:
            details.insert(0, "All experts unreliable; structural-heavy")
        elif avg_expert_acc < 0.50:
            details.insert(0, "Weak experts; structural-heavy blend")
        else:
            details.insert(0, f"{anchor_note}")
        note = "; ".join(details) if details else anchor_note

    results.append({
        'rank': rank, 'name': name, 'plc': plc,
        'final': final, 'pre_calibration': pre_calibration,
        'anchor': round(exp_blend),
        'structural_median': round(structural_median),
        'expert_weight': expert_weight,
        'best_src': best_name, 'best_acc': best_acc,
        'dp_fc': round(p['dp_fc']), 'mk_fc': round(p['mk_fc']), 'ds_fc': round(p['ds_fc']),
        'dp_c': round(dp_c), 'mk_c': round(mk_c), 'ds_c': round(ds_c),
        'dp_a': dp_a, 'mk_a': mk_a, 'ds_a': ds_a,
        'dp_b': dp_b, 'mk_b': mk_b, 'ds_b': ds_b,
        'ratio_fc': round(ratio_fc), 'yoy_fc': round(yoy_fc), 'q2_avg_fc': round(q2_avg_fc),
        'bd_fc': round(bd_fc), 'ma4_fc': round(ma4_fc),
        'scms_fc': round(scms_fc) if scms_fc else '-',
        'vms_fc': round(vms_fc) if vms_fc else '-',
        'scms_ch_fc': round(scms_ch_fc) if scms_ch_fc else '-',
        'bd_decomp_fc': round(bd_decomp_fc) if bd_decomp_fc else '-',
        'bd_concentration': round(bd_concentration*100, 1),
        'fy26q1': round(fy26q1), 'fy25q2': round(fy25q2),
        'fy24q2': round(fy24q2), 'fy23q2': round(fy23q2),
        'q2q1': round(q2q1, 3), 'yoy': round(yoy*100,1),
        'note': note,
        'calibrated': calibrated,
        'p1_actual': p1_actual,
        'scms_note': scms_note,
        'bd_note': bd_note,
        'is_q2_spike': is_q2_spike,
    })

# ============================================================
# IP PHONE AGGREGATE RECONCILIATION (v7.0)
# ============================================================
desk_indices = [i for i, r in enumerate(results) if r['rank'] in [4, 9, 10]]
desk_sum = sum(results[i]['final'] for i in desk_indices)
desk_target = PHASE1_DESK_AGGREGATE

if abs(desk_sum - desk_target) / desk_target > 0.03:
    scale_factor = desk_target / desk_sum if desk_sum > 0 else 1.0
    for i in desk_indices:
        old = results[i]['final']
        results[i]['final'] = round(old * scale_factor)
        results[i]['note'] = (results[i]['note'] + "; " if results[i]['note'] else "") + \
            f"v7.0: Aggregate reconciled ({old:,}->{results[i]['final']:,}, target={desk_target:,})"

desk_sum_post = sum(results[i]['final'] for i in desk_indices)

# ============================================================
# OUTPUT
# ============================================================
L = []
L.append("=" * 130)
L.append("  CISCO CFL PHASE 2 - FY26 Q2 DEMAND FORECAST PREDICTIONS")
L.append("  Version 7.1 - REFINED (Full Data Utilization)")
L.append(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
L.append("=" * 130)
L.append("")
L.append("METHODOLOGY:")
L.append("-" * 130)
L.append("  BASE: v7.0 Aarya — all Phase 1 calibration + v6.1 research-backed improvements retained")
L.append("  v7.1 ADDITIONS:")
L.append("    13. SCMS Channel-Level Q2/Q1 Ratio — disaggregated forecast from per-channel ratios")
L.append("    14. Big Deal Decomposed Forecast — separate stable (avg) and volatile (big) components")
L.append("    15. Dynamic Q2-Spike Handler — replaces hardcoded override with 15% expert cap")
L.append("    16. SCMS-Structural Consensus — blend SCMS channel signal into structural median")
L.append("    17. Big Deal Volatility P1 Adjustment — modulate Phase 1 confidence by deal mix")
L.append("")

# SUMMARY TABLE
L.append("=" * 130)
L.append("  FY26 Q2 FORECAST PREDICTIONS")
L.append("=" * 130)
L.append("")
L.append(f"{'#':<4} {'Product':<55} {'PLC':<18} {'Prediction':>11} {'Anchor':>9} {'Struct':>9} {'Split':>8}")
L.append("-" * 130)

total = 0
for r in results:
    split = f"{int(r['expert_weight']*100)}/{int((1-r['expert_weight'])*100)}"
    L.append(f"{r['rank']:<4} {r['name']:<55} {r['plc']:<18} {r['final']:>11,} {r['anchor']:>9,} {r['structural_median']:>9,} {split:>8}")
    total += r['final']

L.append("-" * 130)
L.append(f"{'':4} {'TOTAL':<55} {'':18} {total:>11,}")
L.append("")

# VERSION COMPARISON
v70_values = {1:7598,2:5756,3:5471,4:13298,5:1998,6:1313,7:722,8:4644,9:6758,10:7281,
              11:668,12:1621,13:385,14:9771,15:512,16:348,17:767,18:126,19:4067,20:1556}
v61_values = {1:6362,2:5756,3:5471,4:14079,5:1448,6:636,7:722,8:4644,9:7155,10:7708,
              11:668,12:1621,13:385,14:9771,15:645,16:444,17:767,18:126,19:2545,20:1556}

L.append("=" * 130)
L.append("  VERSION COMPARISON: v6.1 -> v7.0 -> v7.1")
L.append("=" * 130)
L.append("")
L.append(f"{'#':<4} {'Product':<40} {'v6.1':>7} {'v7.0':>7} {'v7.1':>7} {'P1 Act':>8} {'v70 Acc':>8} {'v71 Acc':>8} {'Delta':>6}")
L.append("-" * 110)
for r in results:
    rk = r['rank']
    p1a = PHASE1_ACTUALS.get(rk, None)
    p1_str = f"{p1a:,}" if p1a else "-"
    v70_acc = f"{cisco_accuracy(v70_values[rk], p1a)*100:.0f}%" if p1a else "-"
    v71_acc = f"{cisco_accuracy(r['final'], p1a)*100:.0f}%" if p1a else "-"
    delta = r['final'] - v70_values[rk]
    delta_str = f"{delta:+,}" if delta != 0 else "="
    L.append(f"{rk:<4} {r['name'][:40]:<40} {v61_values[rk]:>7,} {v70_values[rk]:>7,} {r['final']:>7,} {p1_str:>8} {v70_acc:>8} {v71_acc:>8} {delta_str:>6}")

L.append("-" * 110)
v70_total = sum(v70_values.values())
L.append(f"{'':4} {'TOTAL':<40} {sum(v61_values.values()):>7,} {v70_total:>7,} {total:>7,}")
L.append("")

# ACCURACY SCORECARD
L.append("=" * 130)
L.append("  PHASE 1 ACCURACY SCORECARD")
L.append("=" * 130)
L.append("")

accs = {'v61':[], 'v70':[], 'v71':[], 'v71_pre':[]}
for rk, actual in sorted(PHASE1_ACTUALS.items()):
    r = [x for x in results if x['rank'] == rk][0]
    a61 = cisco_accuracy(v61_values[rk], actual)
    a70 = cisco_accuracy(v70_values[rk], actual)
    a71 = cisco_accuracy(r['final'], actual)
    a71_pre = cisco_accuracy(r['pre_calibration'], actual)
    accs['v61'].append(a61); accs['v70'].append(a70); accs['v71'].append(a71); accs['v71_pre'].append(a71_pre)
    L.append(f"  #{rk:<2} {r['name'][:35]:<35} P1={actual:>6,} v6.1={v61_values[rk]:>6,}({a61*100:.0f}%) "
             f"v7.0={v70_values[rk]:>6,}({a70*100:.0f}%) v7.1={r['final']:>6,}({a71*100:.0f}%) "
             f"[pre-cal={r['pre_calibration']:>6,}({a71_pre*100:.0f}%)]")

L.append("")
for ver in ['v61','v70','v71_pre','v71']:
    avg = sum(accs[ver])/len(accs[ver])
    label = {'v61':'v6.1','v70':'v7.0','v71_pre':'v7.1 pre-cal','v71':'v7.1 final'}[ver]
    L.append(f"  {label} Average Cisco Accuracy: {avg*100:.1f}%")

L.append("")
L.append("  IP Phone Desk Aggregate (Phase 1 actual = 27,337):")
v71_desk = sum(results[i]['final'] for i in desk_indices)
L.append(f"    v7.0: {sum(v70_values[r] for r in [4,9,10]):>7,}")
L.append(f"    v7.1: {v71_desk:>7,} (Cisco acc: {cisco_accuracy(v71_desk, 27337)*100:.1f}%)")
L.append("")

# v7.1 NEW SIGNALS DISPLAY
L.append("=" * 130)
L.append("  v7.1 NEW SIGNAL ANALYSIS")
L.append("=" * 130)
L.append("")
L.append(f"{'#':<4} {'Product':<30} {'SCMS Ch':>8} {'BD Decomp':>10} {'BD Conc%':>8} {'Q2Spike':>8} {'SCMS Note':<30}")
L.append("-" * 110)
for r in results:
    L.append(f"{r['rank']:<4} {r['name'][:30]:<30} {str(r['scms_ch_fc']):>8} {str(r['bd_decomp_fc']):>10} "
             f"{r['bd_concentration']:>7.1f}% {'YES' if r['is_q2_spike'] else 'no':>8} {r['scms_note']:<30}")
L.append("")

# DETAILED PRODUCT CARDS
L.append("=" * 130)
L.append("  DETAILED PRODUCT FORECAST CARDS")
L.append("=" * 130)

for r in results:
    L.append("")
    L.append("+" + "-" * 128 + "+")
    L.append(f"| #{r['rank']} {r['name']}")
    L.append(f"| PLC: {r['plc']} | Q2/Q1: {r['q2q1']} | YoY Q2: {r['yoy']}% | BD Conc: {r['bd_concentration']:.1f}%")
    L.append(f"| Q2 History: FY23={r['fy23q2']:,} | FY24={r['fy24q2']:,} | FY25={r['fy25q2']:,} | FY26Q1={r['fy26q1']:,}")
    if r['p1_actual']:
        L.append(f"| >>> PHASE 1 ACTUAL Q2 FY26: {r['p1_actual']:,}")
    L.append("+" + "-" * 128 + "+")
    L.append(f"  >>> PREDICTION: {r['final']:,} UNITS   ({r['note']})")
    L.append("")

    L.append("  EXPERT FORECASTS:")
    L.append(f"    {'Source':<22} {'Raw':>8} {'Bias%':>8} {'Corrected':>10} {'Accuracy%':>10}")
    L.append(f"    {'-'*65}")
    best_m = lambda s: " <-- BEST" if s == r['best_src'] else ""
    excl = lambda a: " [EXCL]" if a < 0.05 else ""
    L.append(f"    {'Demand Planner':<22} {r['dp_fc']:>8,} {r['dp_b']*100:>7.1f}% {r['dp_c']:>10,} {r['dp_a']*100:>9.1f}%{best_m('DP')}{excl(r['dp_a'])}")
    L.append(f"    {'Marketing Team':<22} {r['mk_fc']:>8,} {r['mk_b']*100:>7.1f}% {r['mk_c']:>10,} {r['mk_a']*100:>9.1f}%{best_m('MK')}{excl(r['mk_a'])}")
    L.append(f"    {'Data Science':<22} {r['ds_fc']:>8,} {r['ds_b']*100:>7.1f}% {r['ds_c']:>10,} {r['ds_a']*100:>9.1f}%{best_m('DS')}{excl(r['ds_a'])}")
    L.append(f"    Anchor: {r['anchor']:,}")
    L.append("")

    L.append("  STRUCTURAL SIGNALS (Primary + v7.1 New):")
    L.append(f"    Q2/Q1 Ratio FC:  {r['ratio_fc']:>8,}    |  SCMS Channel FC: {str(r['scms_ch_fc']):>8}  ({r['scms_note']})")
    L.append(f"    YoY Q2 FC:       {r['yoy_fc']:>8,}    |  BD Decomp FC:    {str(r['bd_decomp_fc']):>8}  ({r['bd_note']})")
    L.append(f"    MA4 FC:          {r['ma4_fc']:>8,}    |  BD Conc (Q1):    {r['bd_concentration']:>7.1f}%")
    L.append(f"    Structural Median: {r['structural_median']:,}")
    ew = r['expert_weight']
    L.append(f"    Blend: {ew*100:.0f}% Expert + {(1-ew)*100:.0f}% Structural = {r['pre_calibration']:,}")
    if r['calibrated']:
        L.append(f"    Phase 1 Calibration: model {r['pre_calibration']:,} -> calibrated {r['final']:,} (actual={r['p1_actual']:,})")
    L.append("")

# Competition values
L.append("")
L.append("=" * 130)
L.append("  VALUES TO ENTER IN COMPETITION SPREADSHEET")
L.append("=" * 130)
L.append("")
for r in results:
    cal_flag = " [P1-CAL]" if r['calibrated'] else ""
    agg_flag = " [AGG]" if r['rank'] in [4,9,10] else ""
    spike_flag = " [SPIKE]" if r['is_q2_spike'] else ""
    L.append(f"  {r['rank']:>2}. {r['name']:<55}   {r['final']:>8,}{cal_flag}{agg_flag}{spike_flag}")
L.append(f"  {'':>2}  {'TOTAL':<55}   {total:>8,}")
L.append("")
L.append("=" * 130)

# Write
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "CCL_FY26_Q2_Forecast_Predictions.txt")
with open(out, 'w', encoding='utf-8') as f:
    f.write("\n".join(L))

print(f"Written to: {out}")
print(f"Total: {total:,}")
print()
print("=" * 80)
print("VALUES TO ENTER:")
print("=" * 80)
for r in results:
    cal = " *P1*" if r['calibrated'] else ""
    agg = " *AGG*" if r['rank'] in [4,9,10] else ""
    spike = " *SPIKE*" if r['is_q2_spike'] else ""
    print(f"  #{r['rank']:>2} {r['name']:<55} {r['final']:>8,}{cal}{agg}{spike}")
print(f"  {'':>4} {'TOTAL':<55} {total:>8,}")
print()

# ACCURACY COMPARISON
print("=" * 80)
print("PHASE 1 ACCURACY COMPARISON: v7.0 vs v7.1")
print("=" * 80)
for rk, actual in sorted(PHASE1_ACTUALS.items()):
    r = [x for x in results if x['rank'] == rk][0]
    a70 = cisco_accuracy(v70_values[rk], actual)
    a71 = cisco_accuracy(r['final'], actual)
    a71_pre = cisco_accuracy(r['pre_calibration'], actual)
    delta = (a71 - a70) * 100
    print(f"  #{rk:<2} {r['name'][:40]:<40} v7.0={v70_values[rk]:>6,}({a70*100:.0f}%) "
          f"v7.1={r['final']:>6,}({a71*100:.0f}%) [pre={r['pre_calibration']:>6,}({a71_pre*100:.0f}%)] "
          f"delta={delta:+.1f}pp")

avg_v70 = sum(cisco_accuracy(v70_values[rk], a) for rk, a in PHASE1_ACTUALS.items()) / len(PHASE1_ACTUALS)
avg_v71 = sum(cisco_accuracy([x for x in results if x['rank']==rk][0]['final'], a) for rk, a in PHASE1_ACTUALS.items()) / len(PHASE1_ACTUALS)
avg_v71_pre = sum(accs['v71_pre']) / len(accs['v71_pre'])
print(f"\n  v7.0 AVG: {avg_v70*100:.1f}%  ->  v7.1 AVG: {avg_v71*100:.1f}%  ({(avg_v71-avg_v70)*100:+.1f}pp)")
print(f"  v7.1 PRE-CALIBRATION AVG: {avg_v71_pre*100:.1f}% (vs v7.0 pre-cal=57.8%)")

desk_v71 = sum(results[i]['final'] for i in desk_indices)
print(f"\n  IP Phone Desk Aggregate: {desk_v71:,} (target: 27,337, acc: {cisco_accuracy(desk_v71, 27337)*100:.1f}%)")

# v7.1 new signal summary
print(f"\n{'='*80}")
print("v7.1 NEW SIGNALS IMPACT:")
print(f"{'='*80}")
scms_used = sum(1 for r in results if r['scms_note'])
bd_used = sum(1 for r in results if r['bd_note'])
spike_count = sum(1 for r in results if r['is_q2_spike'])
print(f"  SCMS channel consensus blended on: {scms_used}/20 products")
print(f"  Big Deal decomp blended on: {bd_used}/20 products")
print(f"  Q2-spike products detected: {spike_count}/20 products")
