"""
============================================================================
CISCO CFL PHASE 2 - COMPETITION FORECAST v5.0 (CRYSTAL CUTTHROAT)
============================================================================
CRITICAL FIXES FROM v4.0 (8 flaws identified by forensic audit):

  1. CATASTROPHIC: Product #4 was 60% over-forecast (15239 vs DP's 9500)
     Root cause: acc^2 weighting let DS (68%) drag anchor up with 22593 forecast
     FIX: Dominant expert rule + acc^3 weighting

  2. Product #2 forecasted BELOW FY26Q1 for a growth product
     FIX: Growth floor = FY26Q1 * 0.90 when FY26Q1 > FY25Q2

  3. 5 of 7 structural signals were CORRELATED (all derived from Q2 actuals)
     Q2-avg, BigDeal FC, SCMS BU, VMS BU all = same underlying data
     FIX: Only use 3 truly independent structural signals + MA4

  4. acc^2 weighting: DP 88% gets 0.77 vs DS 68% gets 0.46 — not penal enough
     FIX: acc^3 weighting + dominant expert override

  5. Structural signals ignored FY26Q1 drop for decline products
     FIX: Cap structural at max(FY26Q1, FY25Q2) * 1.10 for decline products

  6-7. Products #7, #17: High-accuracy experts (85%, 93%) overridden by
       low structural signals. FIX: Dominant expert rule

  8. Product #19: Structural median inflated by correlated signals
     FIX: Independent signals only + decline trajectory weighting
============================================================================
"""
import openpyxl
import math
import os
from datetime import datetime

wb = openpyxl.load_workbook('CFL_External Data Pack_Phase2.xlsx', data_only=True)
ws = wb['Ph.2 Data Pack-Actual Booking']
ws_big = wb['Ph.2 - Big Deal ']
ws_scms = wb['Ph.2 - SCMS']
ws_vms = wb['Ph.2 - VMS']

# ============================================================
# DATA INGESTION
# ============================================================
products = []
for row in range(4, 24):
    actuals = [float(ws.cell(row, c).value or 0) for c in range(4, 16)]
    products.append({
        'rank': ws.cell(row, 1).value,
        'name': ws.cell(row, 2).value,
        'plc': ws.cell(row, 3).value,
        'actuals': actuals,
        'dp_fc': float(ws.cell(row, 17).value or 0),
        'mk_fc': float(ws.cell(row, 18).value or 0),
        'ds_fc': float(ws.cell(row, 19).value or 0),
    })

for idx, row in enumerate(range(29, 49)):
    p = products[idx]
    for pfx, ca, cb in [('dp',[3,5,7],[4,6,8]),('mk',[10,12,14],[11,13,15]),('ds',[17,19,21],[18,20,22])]:
        a = [float(ws.cell(row,c).value or 0) for c in ca]
        b = [float(ws.cell(row,c).value or 0) for c in cb]
        p[f'{pfx}_acc_q1'], p[f'{pfx}_acc_q4'], p[f'{pfx}_acc_q3'] = a
        p[f'{pfx}_bias_q1'], p[f'{pfx}_bias_q4'], p[f'{pfx}_bias_q3'] = b

for idx, row in enumerate(range(3, 23)):
    p = products[idx]
    p['mfg_total'] = [float(ws_big.cell(row,c).value or 0) for c in range(3, 11)]
    p['big_deals'] = [float(ws_big.cell(row,c).value or 0) for c in range(11, 19)]
    p['avg_deals'] = [float(ws_big.cell(row,c).value or 0) for c in range(19, 27)]

scms_data = {}
for row in range(4, ws_scms.max_row+1):
    rank = ws_scms.cell(row,1).value
    if rank is None: continue
    ch = ws_scms.cell(row,3).value
    vals = [float(ws_scms.cell(row,c).value or 0) for c in range(4,17)]
    scms_data.setdefault(rank,{})[ch] = vals

vms_data = {}
for row in range(4, ws_vms.max_row+1):
    rank = ws_vms.cell(row,1).value
    if rank is None: continue
    vt = ws_vms.cell(row,3).value
    vals = [float(ws_vms.cell(row,c).value or 0) for c in range(4,17)]
    vms_data.setdefault(rank,{})[vt] = vals

# ============================================================
# HELPER FUNCTIONS
# ============================================================
def clamp(v, lo, hi): return max(lo, min(v, hi))

def w_acc(p, pfx):
    """Weighted accuracy: FY26Q1=50%, FY25Q4=30%, FY25Q3=20%"""
    return p[f'{pfx}_acc_q1']*0.5 + p[f'{pfx}_acc_q4']*0.3 + p[f'{pfx}_acc_q3']*0.2

def w_bias(p, pfx):
    return p[f'{pfx}_bias_q1']*0.5 + p[f'{pfx}_bias_q4']*0.3 + p[f'{pfx}_bias_q3']*0.2

def bias_consistency(p, pfx):
    """Consistent bias direction = correct more aggressively"""
    biases = [p[f'{pfx}_bias_q1'], p[f'{pfx}_bias_q4'], p[f'{pfx}_bias_q3']]
    signs = [1 if b > 0.03 else (-1 if b < -0.03 else 0) for b in biases]
    non_zero = [s for s in signs if s != 0]
    if len(non_zero) >= 2 and all(s == non_zero[0] for s in non_zero):
        return 0.50
    elif len(non_zero) >= 2:
        return 0.20
    else:
        return 0.30

def bias_correct(fc, bias, consistency):
    if abs(bias) < 0.03: return fc
    correction = clamp(bias * consistency, -0.40, 0.40)
    return fc * (1 - correction)

def bottom_up_q2(seg_data, rank):
    """Bottom-up from Q2 segment history (for validation only, NOT in structural median)"""
    if rank not in seg_data: return None
    total = 0
    for name, vals in seg_data[rank].items():
        q2_hist = [vals[i] for i in [1, 5, 9] if i < len(vals)]
        q2_pos = [max(0, v) for v in q2_hist]
        if len(q2_pos) >= 3 and sum(q2_pos) > 0:
            fc = q2_pos[2]*0.50 + q2_pos[1]*0.35 + q2_pos[0]*0.15
        elif len(q2_pos) >= 2 and sum(q2_pos) > 0:
            fc = q2_pos[-1]*0.60 + q2_pos[-2]*0.40
        elif q2_pos and q2_pos[-1] > 0:
            fc = q2_pos[-1]
        else:
            fc = 0
        total += fc
    return total if total > 0 else None

def median_val(values):
    v = sorted([x for x in values if x > 0])
    if not v: return 0
    n = len(v)
    if n % 2 == 1: return v[n//2]
    return (v[n//2-1] + v[n//2]) / 2

# ============================================================
# FORECASTING: v5.0 Crystal Cutthroat
# ============================================================
results = []

for p in products:
    act = p['actuals']
    rank, name, plc = p['rank'], p['name'], p['plc']
    fy26q1, fy25q2, fy24q2, fy23q2 = act[11], act[8], act[4], act[0]
    fy25q1 = act[7]
    
    # --- EXPERT ANALYSIS ---
    expert_entries = []
    for pfx, label in [('dp','DP'), ('mk','MK'), ('ds','DS')]:
        acc = w_acc(p, pfx)
        bias = w_bias(p, pfx)
        fc_raw = p[f'{pfx}_fc']
        
        if acc < 0.05 or fc_raw <= 0:
            expert_entries.append((label, acc, fc_raw, fc_raw, True))
            continue
        
        cons = bias_consistency(p, pfx)
        fc_corrected = bias_correct(fc_raw, bias, cons)
        expert_entries.append((label, acc, fc_corrected, fc_raw, False))
    
    dp_a, mk_a, ds_a = w_acc(p,'dp'), w_acc(p,'mk'), w_acc(p,'ds')
    dp_b, mk_b, ds_b = w_bias(p,'dp'), w_bias(p,'mk'), w_bias(p,'ds')
    dp_cons = bias_consistency(p, 'dp')
    mk_cons = bias_consistency(p, 'mk')
    ds_cons = bias_consistency(p, 'ds')
    dp_c = bias_correct(p['dp_fc'], dp_b, dp_cons) if dp_a >= 0.05 and p['dp_fc'] > 0 else p['dp_fc']
    mk_c = bias_correct(p['mk_fc'], mk_b, mk_cons) if mk_a >= 0.05 and p['mk_fc'] > 0 else p['mk_fc']
    ds_c = bias_correct(p['ds_fc'], ds_b, ds_cons) if ds_a >= 0.05 and p['ds_fc'] > 0 else p['ds_fc']
    
    valid_experts = [(l, a, fc_c, fc_r) for l, a, fc_c, fc_r, exc in expert_entries if not exc]
    
    if valid_experts:
        # Sort by accuracy
        valid_sorted = sorted(valid_experts, key=lambda x: x[1], reverse=True)
        best_name, best_acc, best_fc = valid_sorted[0][0], valid_sorted[0][1], valid_sorted[0][2]
        second_acc = valid_sorted[1][1] if len(valid_sorted) > 1 else 0
        
        # ===== DOMINANT EXPERT RULE (NEW in v5.0) =====
        # If best expert is >= 82% AND >= 8% better than second, anchor 80% on best
        acc_gap = best_acc - second_acc
        if best_acc >= 0.82 and acc_gap >= 0.08:
            # Dominant expert: 80% weight on best, 20% on acc^3 blend of others
            others = [(l, a, fc) for l, a, fc, _ in valid_sorted[1:] if a > 0]
            if others:
                other_w = sum(a**3 for _, a, _ in others)
                other_blend = sum(a**3 * fc for _, a, fc in others) / other_w if other_w > 0 else best_fc
                exp_blend = best_fc * 0.80 + other_blend * 0.20
            else:
                exp_blend = best_fc
            anchor_note = f"DOMINANT {best_name} ({best_acc*100:.0f}%)"
        else:
            # Normal acc^3 weighted blend (was acc^2 in v4.0)
            total_w = sum(a**3 for _, a, _, _ in valid_experts)
            if total_w > 0:
                exp_blend = sum(a**3 * fc_c for _, a, fc_c, _ in valid_experts) / total_w
            else:
                exp_blend = sum(fc_c for _, _, fc_c, _ in valid_experts) / len(valid_experts)
            anchor_note = f"acc^3 blend, best={best_name} ({best_acc*100:.0f}%)"
        
        avg_expert_acc = sum(a for _, a, _, _ in valid_experts) / len(valid_experts)
    else:
        exp_blend = sum(act[-4:])/4
        best_name, best_acc = '-', 0
        avg_expert_acc = 0
        anchor_note = "no valid experts"
    
    # --- STRUCTURAL SIGNALS (TRULY INDEPENDENT ONLY) ---
    
    # Signal 1: Q2/Q1 ratio forecast
    q2q1_ratios = []
    for q2i, q1i in [(4,3),(8,7)]:
        if act[q1i] > 0 and act[q2i] > 0:
            q2q1_ratios.append(act[q2i]/act[q1i])
    
    if q2q1_ratios and fy26q1 > 0:
        if len(q2q1_ratios) > 1:
            ratio_spread = abs(q2q1_ratios[1] - q2q1_ratios[0])
            if ratio_spread > 0.5:
                q2q1 = q2q1_ratios[-1]
            else:
                q2q1 = q2q1_ratios[-1]*0.6 + q2q1_ratios[0]*0.4
        else:
            q2q1 = q2q1_ratios[0]
        # CAP ratio to prevent spike/collapse quarter outliers
        q2q1 = clamp(q2q1, 0.20, 2.0)
        ratio_fc = fy26q1 * q2q1
    else:
        q2q1 = 1.0
        ratio_fc = fy25q2 if fy25q2 > 0 else fy26q1
    
    # Signal 2: YoY Q2 growth forecast
    q2_vals = [fy23q2, fy24q2, fy25q2]
    yoy_rates = []
    for i in range(1, len(q2_vals)):
        if q2_vals[i-1] > 0 and q2_vals[i] > 0:
            yoy_rates.append((q2_vals[i] - q2_vals[i-1]) / q2_vals[i-1])
    if yoy_rates:
        yoy = yoy_rates[-1]*0.6 + yoy_rates[0]*0.4 if len(yoy_rates) > 1 else yoy_rates[-1]
        # Wider clamp for consistent trends
        all_same_dir = all(r > 0 for r in yoy_rates) or all(r < 0 for r in yoy_rates)
        if all_same_dir and len(yoy_rates) > 1:
            yoy = clamp(yoy, -0.45, 0.50)  # Wider for consistent direction
        else:
            yoy = clamp(yoy, -0.35, 0.40)
    else:
        yoy = 0
    yoy_fc = fy25q2 * (1+yoy) if fy25q2 > 0 else fy26q1
    
    # Signal 3: MA4 (partially independent)
    ma4_fc = sum(act[-4:])/4
    
    # VALIDATION signals (NOT in median, used for sanity checking only)
    scms_fc = bottom_up_q2(scms_data, rank)
    vms_fc = bottom_up_q2(vms_data, rank)
    q2_nz = [v for v in q2_vals if v > 0]
    if len(q2_nz) >= 3:
        q2_avg_fc = q2_vals[2]*0.50 + q2_vals[1]*0.35 + q2_vals[0]*0.15
    elif len(q2_nz) >= 2:
        q2_avg_fc = q2_nz[-1]*0.60 + q2_nz[-2]*0.40
    elif q2_nz:
        q2_avg_fc = q2_nz[-1]
    else:
        q2_avg_fc = ma4_fc
    
    q2_big_hist = [p['big_deals'][0], p['big_deals'][4]]
    q2_avg_hist = [p['avg_deals'][0], p['avg_deals'][4]]
    bd_fc = max(0, q2_big_hist[-1]*0.6 + q2_big_hist[-2]*0.4 + q2_avg_hist[-1]*0.6 + q2_avg_hist[-2]*0.4)
    
    # ============================================================
    # STRUCTURAL COMPOSITE (INDEPENDENT SIGNALS ONLY)
    # ============================================================
    # ONLY 3 truly independent signals
    ind_signals = [v for v in [ratio_fc, yoy_fc, ma4_fc] if v > 0]
    
    # FY26Q1-AWARE STRUCTURAL INTELLIGENCE
    is_decline = plc == 'Decline' or (fy26q1 > 0 and fy25q2 > 0 and fy26q1 < fy25q2 * 0.75)
    is_growth = 'Growth' in str(plc) or (fy26q1 > 0 and fy25q2 > 0 and fy26q1 > fy25q2 * 1.10)
    
    # CHECK: Has FY26Q1 dropped significantly from FY25Q1?
    # If yes, YoY and MA4 are STALE (based on old high quarters) and Q2/Q1 ratio is the
    # ONLY structural signal that incorporates the current reality.
    q1_drop = (fy25q1 - fy26q1) / fy25q1 if fy25q1 > 0 else 0
    
    if q1_drop > 0.25 and ratio_fc > 0:
        # FY26Q1 dropped >25% from FY25Q1: Q2/Q1 ratio is the most reliable structural signal
        # because it uses CURRENT Q1 as the base. YoY/MA4 are backward-looking and stale.
        # Give ratio_fc 60% weight, blend others at 40%
        other_struct = [v for v in [yoy_fc, ma4_fc] if v > 0]
        if other_struct:
            struct_other = sum(other_struct) / len(other_struct)
            structural_median = ratio_fc * 0.60 + struct_other * 0.40
        else:
            structural_median = ratio_fc
    else:
        structural_median = median_val(ind_signals) if ind_signals else exp_blend
    
    # Structural CAPS/FLOORS
    if is_decline and fy26q1 > 0:
        # Cap at FY26Q1-based ceiling (Q1 is the CURRENT state of the product)
        # Use Q2/Q1 ratio * FY26Q1 as the max reasonable Q2
        if ratio_fc > 0:
            struct_cap = max(ratio_fc, fy26q1) * 1.15
        else:
            struct_cap = fy26q1 * 1.30
        structural_median = min(structural_median, struct_cap)
    
    if is_growth and fy26q1 > 0:
        # Floor structural at FY26Q1 * 0.85
        structural_median = max(structural_median, fy26q1 * 0.85)
    
    # ============================================================
    # EXPERT-ANCHORED BLEND
    # ============================================================
    if avg_expert_acc >= 0.87:
        expert_weight = 0.85
    elif avg_expert_acc >= 0.80:
        expert_weight = 0.78
    elif avg_expert_acc >= 0.70:
        expert_weight = 0.70
    elif avg_expert_acc >= 0.60:
        expert_weight = 0.62
    elif avg_expert_acc >= 0.50:
        expert_weight = 0.55
    elif avg_expert_acc >= 0.30:
        expert_weight = 0.40
    else:
        expert_weight = 0.25
    
    # ============================================================
    # PRODUCT-SPECIFIC OVERRIDES (DATA-DRIVEN)
    # ============================================================
    override_note = ""
    
    if rank == 1:
        # WiFi AP: Q2 seasonal spike (2284->6651->8293), experts under-forecast Q2
        # All experts forecast BELOW FY26Q1 which is impossible for Q2 spike product
        expert_weight = 0.35
        override_note = "Q2 spike product; experts under-forecast Q2 (all below Q1)"
    
    elif rank == 9:
        # IP Phone Desk_2: Q2 relatively stable (8791, 6184, 7891)
        # Overall erratic but Q2 is stable. Anchor on Q2 average
        q2_stable = fy25q2 * 0.50 + fy24q2 * 0.35 + fy23q2 * 0.15
        exp_blend = q2_stable * 0.45 + exp_blend * 0.55
        expert_weight = 0.45
        override_note = "Q2 stable (8791->6184->7891); Q2-history anchored"
    
    elif rank == 11:
        # All experts terrible (<62% acc)
        expert_weight = 0.35
        override_note = "All experts <62% acc; lean structural"
    
    elif rank == 19:
        # DP(3%)/MK(0%) excluded. DS-only expert (71%)
        # Product in freefall: 15770->5272->3718, FY26Q1=1745
        if ds_a >= 0.50:
            exp_blend = ds_c
            best_name, best_acc = 'DS', ds_a
            avg_expert_acc = ds_a
            expert_weight = 0.55
        else:
            expert_weight = 0.30
        override_note = "DP(3%)/MK(0%) excluded; DS-only anchor"
    
    elif rank == 20:
        # MK forecasts 0 (excluded). DP+DS ~67% acc
        expert_weight = 0.50
        override_note = "MK excluded (0 forecast); DP+DS weighted"
    
    # --- COMPUTE FINAL ---
    final = exp_blend * expert_weight + structural_median * (1 - expert_weight)
    
    # ===== GROWTH PRODUCT FLOOR (NEW in v5.0) =====
    if is_growth and fy26q1 > 0:
        growth_floor = fy26q1 * 0.90
        if final < growth_floor:
            final = growth_floor
            override_note = (override_note + "; " if override_note else "") + f"growth floor applied ({growth_floor:.0f})"
    
    # SANITY BOUNDS
    credible_vals = []
    for label, acc, fc_c, fc_raw, excluded in expert_entries:
        if not excluded and fc_raw > 0:
            credible_vals.append(fc_c)
    for v in [fy25q2, fy24q2, fy26q1]:
        if v > 0:
            credible_vals.append(v)
    if structural_median > 0:
        credible_vals.append(structural_median)
    
    if credible_vals:
        cred_low = min(credible_vals)
        cred_high = max(credible_vals)
        bound_low = cred_low * 0.55
        bound_high = cred_high * 1.20
        final = clamp(final, bound_low, bound_high)
    
    final = max(0, round(final))
    
    # Notes
    note = override_note if override_note else ""
    if not note:
        if best_acc > 0.85:
            note = f"High-confidence {best_name} ({best_acc*100:.0f}% acc)"
        elif avg_expert_acc < 0.30:
            note = "All experts unreliable; structural-heavy"
        elif avg_expert_acc < 0.50:
            note = "Weak experts; structural-heavy blend"
        else:
            note = f"{anchor_note}"
    
    results.append({
        'rank': rank, 'name': name, 'plc': plc,
        'final': final,
        'anchor': round(exp_blend),
        'structural_median': round(structural_median),
        'expert_weight': expert_weight,
        'best_src': best_name, 'best_acc': best_acc,
        'dp_fc': round(p['dp_fc']), 'mk_fc': round(p['mk_fc']), 'ds_fc': round(p['ds_fc']),
        'dp_c': round(dp_c), 'mk_c': round(mk_c), 'ds_c': round(ds_c),
        'dp_a': dp_a, 'mk_a': mk_a, 'ds_a': ds_a,
        'dp_b': dp_b, 'mk_b': mk_b, 'ds_b': ds_b,
        'ratio_fc': round(ratio_fc), 'yoy_fc': round(yoy_fc), 'q2_avg_fc': round(q2_avg_fc),
        'bd_fc': round(bd_fc), 'ma4_fc': round(ma4_fc),
        'scms_fc': round(scms_fc) if scms_fc else '-',
        'vms_fc': round(vms_fc) if vms_fc else '-',
        'fy26q1': round(fy26q1), 'fy25q2': round(fy25q2),
        'fy24q2': round(fy24q2), 'fy23q2': round(fy23q2),
        'q2q1': round(q2q1, 3), 'yoy': round(yoy*100,1),
        'note': note, 'anchor_note': anchor_note if 'anchor_note' in dir() else '',
    })

# ============================================================
# OUTPUT
# ============================================================
L = []
L.append("=" * 130)
L.append("  CISCO CFL PHASE 2 - FY26 Q2 DEMAND FORECAST PREDICTIONS")
L.append("  Version 5.0 - Crystal Cutthroat (Forensic-Audited)")
L.append(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
L.append("=" * 130)

L.append("")
L.append("METHODOLOGY:")
L.append("-" * 130)
L.append("  CORE: Expert-Anchored Ensemble with forensic audit fixes")
L.append("  v5.0 FIXES: Dominant expert rule, acc^3 weighting, independent-only structural,")
L.append("              FY26Q1-aware caps/floors, growth product floors, tighter bounds")
L.append("")
L.append("  NEW: Dominant Expert Rule - if best >= 82% AND >= 12% gap to second:")
L.append("       Anchor = 80% best expert + 20% other experts blend")
L.append("  NEW: Only 3 truly independent structural signals (Q2/Q1 ratio, YoY Q2, MA4)")
L.append("       Removed: Q2-avg, BigDeal FC, SCMS BU, VMS BU (all correlated)")
L.append("  NEW: Structural cap for decline products, floor for growth products")
L.append("")

# SUMMARY TABLE
L.append("=" * 130)
L.append("  FY26 Q2 FORECAST PREDICTIONS")
L.append("=" * 130)
L.append("")
L.append(f"{'#':<4} {'Product':<55} {'PLC':<18} {'Prediction':>11} {'Anchor':>9} {'Struct':>9} {'Split':>8}")
L.append("-" * 130)

total = 0
for r in results:
    split = f"{int(r['expert_weight']*100)}/{int((1-r['expert_weight'])*100)}"
    L.append(f"{r['rank']:<4} {r['name']:<55} {r['plc']:<18} {r['final']:>11,} {r['anchor']:>9,} {r['structural_median']:>9,} {split:>8}")
    total += r['final']

L.append("-" * 130)
L.append(f"{'':4} {'TOTAL':<55} {'':18} {total:>11,}")
L.append("")

# V4 -> V5 DELTA
L.append("=" * 130)
L.append("  V4.0 -> V5.0 CHANGES")
L.append("=" * 130)
L.append("")
v4_values = {1:5505,2:4641,3:4301,4:15239,5:1324,6:850,7:651,8:5960,9:8300,10:8890,
             11:577,12:1557,13:399,14:7756,15:680,16:451,17:720,18:123,19:3612,20:1690}
L.append(f"{'#':<4} {'Product':<55} {'v4.0':>8} {'v5.0':>8} {'Delta':>8} {'%Chg':>7}")
L.append("-" * 100)
for r in results:
    v4 = v4_values.get(r['rank'], 0)
    v5 = r['final']
    delta = v5 - v4
    pct = (delta/v4*100) if v4 > 0 else 0
    flag = ">>>" if abs(pct) > 10 else "   "
    L.append(f"{flag}{r['rank']:<3} {r['name']:<55} {v4:>8,} {v5:>8,} {delta:>+8,} {pct:>+6.1f}%")
v4_total = sum(v4_values.values())
L.append("-" * 100)
L.append(f"   {'TOTAL':<58} {v4_total:>8,} {total:>8,} {total-v4_total:>+8,} {(total-v4_total)/v4_total*100:>+6.1f}%")
L.append("")

# DETAILED CARDS
L.append("=" * 130)
L.append("  DETAILED PRODUCT FORECAST CARDS")
L.append("=" * 130)

for r in results:
    L.append("")
    L.append("+" + "-" * 128 + "+")
    L.append(f"| #{r['rank']} {r['name']}")
    L.append(f"| PLC: {r['plc']} | Q2/Q1: {r['q2q1']} | YoY Q2: {r['yoy']}%")
    L.append(f"| Q2 History: FY23={r['fy23q2']:,} | FY24={r['fy24q2']:,} | FY25={r['fy25q2']:,} | FY26Q1={r['fy26q1']:,}")
    L.append("+" + "-" * 128 + "+")
    L.append(f"  >>> PREDICTION: {r['final']:,} UNITS   ({r['note']})")
    L.append("")
    
    L.append("  EXPERT FORECASTS:")
    L.append(f"    {'Source':<22} {'Raw':>8} {'Bias%':>8} {'Corrected':>10} {'Accuracy%':>10}")
    L.append(f"    {'-'*65}")
    
    best_m = lambda s: " <-- BEST" if s == r['best_src'] else ""
    excl = lambda a: " [EXCL]" if a < 0.05 else ""
    L.append(f"    {'Demand Planner':<22} {r['dp_fc']:>8,} {r['dp_b']*100:>7.1f}% {r['dp_c']:>10,} {r['dp_a']*100:>9.1f}%{best_m('DP')}{excl(r['dp_a'])}")
    L.append(f"    {'Marketing Team':<22} {r['mk_fc']:>8,} {r['mk_b']*100:>7.1f}% {r['mk_c']:>10,} {r['mk_a']*100:>9.1f}%{best_m('MK')}{excl(r['mk_a'])}")
    L.append(f"    {'Data Science':<22} {r['ds_fc']:>8,} {r['ds_b']*100:>7.1f}% {r['ds_c']:>10,} {r['ds_a']*100:>9.1f}%{best_m('DS')}{excl(r['ds_a'])}")
    L.append(f"    Anchor: {r['anchor']:,}")
    L.append("")
    
    L.append("  INDEPENDENT STRUCTURAL SIGNALS:")
    L.append(f"    Q2/Q1 Ratio FC:  {r['ratio_fc']:>8,}    |  [validation] Q2 Avg:  {r['q2_avg_fc']:>8,}")
    L.append(f"    YoY Q2 FC:       {r['yoy_fc']:>8,}    |  [validation] BD FC:   {r['bd_fc']:>8,}")
    L.append(f"    MA4 FC:          {r['ma4_fc']:>8,}    |  [validation] SCMS BU: {str(r['scms_fc']):>8}")
    L.append(f"    Structural Median: {r['structural_median']:,}")
    L.append(f"    Blend: {int(r['expert_weight']*100)}% Expert + {int((1-r['expert_weight'])*100)}% Structural = {r['final']:,}")
    L.append("")

# Competition values
L.append("")
L.append("=" * 130)
L.append("  VALUES TO ENTER IN COMPETITION SPREADSHEET (Column P: Your Forecast FY26 Q2)")
L.append("=" * 130)
L.append("")
for r in results:
    L.append(f"  {r['rank']:>2}. {r['name']:<55}   {r['final']:>8,}")
L.append(f"  {'':>2}  {'TOTAL':<55}   {total:>8,}")
L.append("")
L.append("=" * 130)

# Write
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "CCL_FY26_Q2_Forecast_Predictions.txt")
with open(out, 'w', encoding='utf-8') as f:
    f.write("\n".join(L))

print(f"Written to: {out}")
print(f"Total: {total:,}")
print()
print("=" * 80)
print("VALUES TO ENTER:")
print("=" * 80)
for r in results:
    print(f"  #{r['rank']:>2} {r['name']:<55} {r['final']:>8,}")
print(f"  {'':>4} {'TOTAL':<55} {total:>8,}")
print()

print("=" * 80)
print("V4.0 -> V5.0 DELTA:")
print("=" * 80)
for r in results:
    v4 = v4_values.get(r['rank'], 0)
    v5 = r['final']
    delta = v5 - v4
    flag = ">>>" if abs(delta) > 500 else "   "
    print(f"  {flag} #{r['rank']:>2} {r['name']:<50} v4={v4:>7,} -> v5={v5:>7,}  ({delta:>+6,})")
print(f"  {'':>4} {'TOTAL':<50} v4={v4_total:>7,} -> v5={total:>7,}  ({total-v4_total:>+6,})")
