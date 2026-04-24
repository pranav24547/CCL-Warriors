import openpyxl

wb = openpyxl.load_workbook('Aarya_v7_full_context_here/CFL_External Data Pack_Phase2.xlsx', data_only=True)
print('=== ALL SHEET NAMES ===')
for name in wb.sheetnames:
    ws = wb[name]
    print(f'  Sheet: "{name}"')
    print(f'    Max row: {ws.max_row}, Max col: {ws.max_column}')
    for r in range(1, min(6, ws.max_row+1)):
        row_data = []
        for c in range(1, min(ws.max_column+1, 25)):
            v = ws.cell(r, c).value
            if v is not None:
                row_data.append(f'C{c}={v}')
        if row_data:
            print(f'    Row {r}: {row_data[:15]}')
    print()
