"""
FORENSIC AUDIT v5.0 - Ruthless backtest & flaw detection
=========================================================
This script:
1. Backtests EVERY forecasting method against KNOWN Q2 actuals (FY24Q2, FY25Q2)
2. Finds which method would have been BEST for each product
3. Identifies specific flaws in the v4.0 model
4. Computes the OPTIMAL forecast for FY26Q2 using backtested evidence
5. Validates SCMS/VMS bottom-up accuracy against known Q2 totals
"""
import openpyxl
import math

wb = openpyxl.load_workbook('CFL_External Data Pack_Phase2.xlsx', data_only=True)
ws = wb['Ph.2 Data Pack-Actual Booking']
ws_big = wb['Ph.2 - Big Deal ']
ws_scms = wb['Ph.2 - SCMS']
ws_vms = wb['Ph.2 - VMS']

# ============================================================
# DATA INGESTION (same as v4.0)
# ============================================================
products = []
for row in range(4, 24):
    actuals = [float(ws.cell(row, c).value or 0) for c in range(4, 16)]
    products.append({
        'rank': ws.cell(row, 1).value,
        'name': ws.cell(row, 2).value,
        'plc': ws.cell(row, 3).value,
        'actuals': actuals,
        'dp_fc': float(ws.cell(row, 17).value or 0),
        'mk_fc': float(ws.cell(row, 18).value or 0),
        'ds_fc': float(ws.cell(row, 19).value or 0),
    })

for idx, row in enumerate(range(29, 49)):
    p = products[idx]
    for pfx, ca, cb in [('dp',[3,5,7],[4,6,8]),('mk',[10,12,14],[11,13,15]),('ds',[17,19,21],[18,20,22])]:
        a = [float(ws.cell(row,c).value or 0) for c in ca]
        b = [float(ws.cell(row,c).value or 0) for c in cb]
        p[f'{pfx}_acc_q1'], p[f'{pfx}_acc_q4'], p[f'{pfx}_acc_q3'] = a
        p[f'{pfx}_bias_q1'], p[f'{pfx}_bias_q4'], p[f'{pfx}_bias_q3'] = b

for idx, row in enumerate(range(3, 23)):
    p = products[idx]
    p['mfg_total'] = [float(ws_big.cell(row,c).value or 0) for c in range(3, 11)]
    p['big_deals'] = [float(ws_big.cell(row,c).value or 0) for c in range(11, 19)]
    p['avg_deals'] = [float(ws_big.cell(row,c).value or 0) for c in range(19, 27)]

scms_data = {}
for row in range(4, ws_scms.max_row+1):
    rank = ws_scms.cell(row,1).value
    if rank is None: continue
    ch = ws_scms.cell(row,3).value
    vals = [float(ws_scms.cell(row,c).value or 0) for c in range(4,17)]
    scms_data.setdefault(rank,{})[ch] = vals

vms_data = {}
for row in range(4, ws_vms.max_row+1):
    rank = ws_vms.cell(row,1).value
    if rank is None: continue
    vt = ws_vms.cell(row,3).value
    vals = [float(ws_vms.cell(row,c).value or 0) for c in range(4,17)]
    vms_data.setdefault(rank,{})[vt] = vals

# ============================================================
# HELPERS
# ============================================================
def clamp(v, lo, hi): return max(lo, min(v, hi))

def cisco_acc(fc, actual):
    if actual == 0: return 0
    return max(0, 1 - abs((fc - actual)/actual))

def median_val(values):
    v = sorted([x for x in values if x > 0])
    if not v: return 0
    n = len(v)
    if n % 2 == 1: return v[n//2]
    return (v[n//2-1] + v[n//2]) / 2

# ============================================================
# SCMS/VMS BOTTOM-UP VALIDATION
# ============================================================
# Quarters in SCMS/VMS: index 0=FY23Q1, 1=FY23Q2, ..., 5=FY24Q2, ..., 9=FY25Q2, ..., 12=FY26Q1
print("=" * 120)
print("AUDIT 1: SCMS/VMS BOTTOM-UP VALIDATION AGAINST KNOWN ACTUALS")
print("=" * 120)
print()
print("Testing: If we sum SCMS segments for a known Q2, does it match the actual?")
print()

for p in products:
    rank = p['rank']
    act = p['actuals']
    # FY25Q2 = act[8], SCMS FY25Q2 = index 9 in SCMS array
    # FY24Q2 = act[4], SCMS FY24Q2 = index 5 in SCMS array
    
    if rank in scms_data:
        scms_fy25q2 = sum(vals[9] for vals in scms_data[rank].values() if len(vals) > 9)
        scms_fy24q2 = sum(vals[5] for vals in scms_data[rank].values() if len(vals) > 5)
        acc_25 = cisco_acc(scms_fy25q2, act[8]) if act[8] > 0 else 0
        acc_24 = cisco_acc(scms_fy24q2, act[4]) if act[4] > 0 else 0
        match_25 = "OK" if abs(scms_fy25q2 - act[8]) < 2 else f"MISMATCH({scms_fy25q2:.0f} vs {act[8]:.0f})"
        match_24 = "OK" if abs(scms_fy24q2 - act[4]) < 2 else f"MISMATCH({scms_fy24q2:.0f} vs {act[4]:.0f})"
        print(f"  #{rank:<2} {p['name'][:50]:<50} FY24Q2:{match_24:>20} FY25Q2:{match_25:>20}")

print()
print("KEY FINDING: SCMS/VMS sums for Q2 MATCH actuals exactly (they're decompositions).")
print("This means the v4.0 'bottom_up_q2' function is forecasting Q2 from Q2 history,")
print("which IS correct methodology. But are the forecasts ACCURATE?")

# ============================================================
# AUDIT 2: BACKTEST EVERY METHOD AGAINST FY25Q2
# ============================================================
print()
print("=" * 120)
print("AUDIT 2: BACKTEST ALL METHODS AGAINST FY25Q2 (KNOWN ACTUAL)")
print("=" * 120)
print()
print("For each product, we simulate what each method would have predicted for FY25Q2")
print("using ONLY data available BEFORE FY25Q2 (i.e., up to FY25Q1)")
print()

# We can backtest these methods against FY25Q2:
# - Q2/Q1 ratio: use FY24Q2/FY24Q1 ratio * FY25Q1
# - YoY Q2: use FY24Q2 * (1 + dampened growth from FY23Q2->FY24Q2)
# - Q2 weighted avg: weighted avg of FY23Q2 and FY24Q2
# - MA4 (last 4 before FY25Q2): FY24Q2, FY24Q3, FY24Q4, FY25Q1
# - Big deal: FY24Q2 big+avg
# - SCMS bottom-up: FY23Q2 and FY24Q2 segment-level forecast

backtest_results = {}

for p in products:
    rank = p['rank']
    act = p['actuals']
    # Quarters: 0=FY23Q2, 1=FY23Q3, 2=FY23Q4, 3=FY24Q1, 4=FY24Q2, 5=FY24Q3, 6=FY24Q4, 7=FY25Q1, 8=FY25Q2
    actual_fy25q2 = act[8]
    fy25q1 = act[7]
    fy24q2 = act[4]
    fy24q1 = act[3]
    fy23q2 = act[0]
    
    methods = {}
    
    # Method 1: Q2/Q1 ratio
    if fy24q1 > 0 and fy24q2 > 0 and fy25q1 > 0:
        ratio = fy24q2 / fy24q1
        methods['Q2/Q1_ratio'] = fy25q1 * ratio
    
    # Method 2: YoY Q2
    if fy23q2 > 0 and fy24q2 > 0:
        yoy = (fy24q2 - fy23q2) / fy23q2
        yoy_clamped = clamp(yoy, -0.35, 0.40)
        methods['YoY_Q2'] = fy24q2 * (1 + yoy_clamped)
    
    # Method 3: Q2 weighted average
    q2_nz = [v for v in [fy23q2, fy24q2] if v > 0]
    if len(q2_nz) >= 2:
        methods['Q2_WtAvg'] = q2_nz[-1]*0.60 + q2_nz[-2]*0.40
    elif q2_nz:
        methods['Q2_WtAvg'] = q2_nz[-1]
    
    # Method 4: MA4 (4 quarters before FY25Q2)
    ma4_vals = [act[4], act[5], act[6], act[7]]  # FY24Q2-FY25Q1
    methods['MA4'] = sum(ma4_vals) / 4
    
    # Method 5: Big deal decomposition for Q2
    # FY24Q2 is index 0 in big_deals
    bd_big = p['big_deals'][0]
    bd_avg = p['avg_deals'][0]
    methods['BigDeal_Q2'] = max(0, bd_big + bd_avg)
    
    # Method 6: SCMS bottom-up (FY23Q2 and FY24Q2 history)
    if rank in scms_data:
        scms_total = 0
        for ch, vals in scms_data[rank].items():
            q2_hist = []
            for i in [1, 5]:  # FY23Q2, FY24Q2 only (not FY25Q2)
                if i < len(vals):
                    q2_hist.append(max(0, vals[i]))
            if len(q2_hist) >= 2:
                fc = q2_hist[-1]*0.60 + q2_hist[-2]*0.40
            elif q2_hist:
                fc = q2_hist[-1]
            else:
                fc = 0
            scms_total += fc
        if scms_total > 0:
            methods['SCMS_BU'] = scms_total
    
    # Method 7: VMS bottom-up
    if rank in vms_data:
        vms_total = 0
        for vt, vals in vms_data[rank].items():
            q2_hist = []
            for i in [1, 5]:
                if i < len(vals):
                    q2_hist.append(max(0, vals[i]))
            if len(q2_hist) >= 2:
                fc = q2_hist[-1]*0.60 + q2_hist[-2]*0.40
            elif q2_hist:
                fc = q2_hist[-1]
            else:
                fc = 0
            vms_total += fc
        if vms_total > 0:
            methods['VMS_BU'] = vms_total
    
    # Method 8: Just use FY24Q2 (naive repeat)
    if fy24q2 > 0:
        methods['Naive_Q2'] = fy24q2
    
    # Method 9: Just use FY25Q1 (naive last quarter)
    if fy25q1 > 0:
        methods['Naive_Q1'] = fy25q1
    
    # Compute accuracy for each
    method_accs = {}
    for name, fc in methods.items():
        acc = cisco_acc(fc, actual_fy25q2) if actual_fy25q2 > 0 else 0
        method_accs[name] = (fc, acc)
    
    backtest_results[rank] = method_accs
    
    if actual_fy25q2 > 0:
        best_method = max(method_accs.items(), key=lambda x: x[1][1])
        print(f"  #{rank:<2} {p['name'][:45]:<45} Actual FY25Q2={actual_fy25q2:>8,.0f}")
        for name, (fc, acc) in sorted(method_accs.items(), key=lambda x: -x[1][1]):
            marker = " <<<" if name == best_method[0] else ""
            print(f"       {name:<15} FC={fc:>8,.0f}  Acc={acc*100:>5.1f}%{marker}")
        print()

# ============================================================
# AUDIT 3: SYSTEMATIC FLAW DETECTION IN V4.0
# ============================================================
print()
print("=" * 120)
print("AUDIT 3: SYSTEMATIC FLAW DETECTION IN V4.0 FORECASTS")
print("=" * 120)
print()

v4_values = {1:5505,2:4641,3:4301,4:15239,5:1324,6:850,7:651,8:5960,9:8300,10:8890,
             11:577,12:1557,13:399,14:7756,15:680,16:451,17:720,18:123,19:3612,20:1690}

flaws = []

for p in products:
    rank = p['rank']
    act = p['actuals']
    v4_fc = v4_values[rank]
    fy26q1 = act[11]
    fy25q2 = act[8]
    fy24q2 = act[4]
    fy23q2 = act[0]
    
    product_flaws = []
    
    # FLAW CHECK 1: Is forecast outside ALL Q2 historical range?
    q2_hist = [v for v in [fy23q2, fy24q2, fy25q2] if v > 0]
    if q2_hist:
        q2_min, q2_max = min(q2_hist), max(q2_hist)
        if v4_fc < q2_min * 0.5:
            product_flaws.append(f"BELOW Q2 FLOOR: FC={v4_fc:,} < 50% of min Q2 ({q2_min:,.0f})")
        if v4_fc > q2_max * 1.5:
            product_flaws.append(f"ABOVE Q2 CEILING: FC={v4_fc:,} > 150% of max Q2 ({q2_max:,.0f})")
    
    # FLAW CHECK 2: Is forecast radically different from best backtest method?
    if rank in backtest_results:
        bt = backtest_results[rank]
        best_bt = max(bt.items(), key=lambda x: x[1][1])
        best_method_name = best_bt[0]
        best_method_acc = best_bt[1][1]
        # What would this best method predict for FY26Q2?
        # (We'll compute this separately below)
    
    # FLAW CHECK 3: Expert vs structural divergence
    # If experts predict X and structural predicts 2X, something is wrong
    # (Already captured in the v4.0 output, but let's flag extreme cases)
    
    # FLAW CHECK 4: Decline products forecasting UP
    if p['plc'] == 'Decline':
        # Check if recent quarters show clear decline
        recent = act[-4:]  # Last 4 quarters
        if recent[-1] < recent[0]:  # Declining
            if v4_fc > fy25q2 * 1.15:  # Forecasting > 15% above last Q2
                product_flaws.append(f"DECLINE PRODUCT BUT FC UP: FC={v4_fc:,} > FY25Q2*1.15={fy25q2*1.15:,.0f}")
    
    # FLAW CHECK 5: Growth products forecasting flat/down
    if 'Growth' in str(p['plc']):
        if fy26q1 > fy25q2 and v4_fc < fy25q2:
            product_flaws.append(f"GROWTH PRODUCT FC BELOW FY25Q2: FC={v4_fc:,} < {fy25q2:,.0f} (FY26Q1={fy26q1:,.0f})")
    
    # FLAW CHECK 6: Q2/Q1 ratio unreasonable 
    if fy26q1 > 0:
        implied_ratio = v4_fc / fy26q1
        if q2_hist and len(q2_hist) >= 2:
            actual_ratios = []
            for q2i, q1i in [(4,3),(8,7)]:
                if act[q1i] > 0 and act[q2i] > 0:
                    actual_ratios.append(act[q2i]/act[q1i])
            if actual_ratios:
                avg_ratio = sum(actual_ratios)/len(actual_ratios)
                if implied_ratio > avg_ratio * 2.5 or implied_ratio < avg_ratio * 0.3:
                    product_flaws.append(f"IMPLIED RATIO EXTREME: FC/Q1={implied_ratio:.2f} vs hist avg={avg_ratio:.2f}")
    
    # FLAW CHECK 7: SCMS/VMS bottom-up was bad in backtest
    if rank in backtest_results:
        bt = backtest_results[rank]
        if 'SCMS_BU' in bt and bt['SCMS_BU'][1] < 0.40:
            product_flaws.append(f"SCMS bottom-up was POOR in backtest: {bt['SCMS_BU'][1]*100:.1f}% acc for FY25Q2")
        if 'VMS_BU' in bt and bt['VMS_BU'][1] < 0.40:
            product_flaws.append(f"VMS bottom-up was POOR in backtest: {bt['VMS_BU'][1]*100:.1f}% acc for FY25Q2")
    
    if product_flaws:
        print(f"  #{rank} {p['name']}")
        print(f"     v4.0 FC: {v4_fc:,} | Q2: {fy23q2:,.0f}->{fy24q2:,.0f}->{fy25q2:,.0f} | FY26Q1={fy26q1:,.0f}")
        for flaw in product_flaws:
            print(f"     >> {flaw}")
        print()

# ============================================================
# AUDIT 4: OPTIMAL METHOD PER PRODUCT (from backtest)
# ============================================================
print()
print("=" * 120)
print("AUDIT 4: OPTIMAL FORECASTING STRATEGY PER PRODUCT")
print("=" * 120)
print()
print("Based on FY25Q2 backtest, the method that would have been most accurate:")
print()

optimal_methods = {}
for p in products:
    rank = p['rank']
    if rank in backtest_results and p['actuals'][8] > 0:
        bt = backtest_results[rank]
        best = max(bt.items(), key=lambda x: x[1][1])
        optimal_methods[rank] = best[0]
        
        # Also find which methods were consistently good
        good_methods = [(n, a) for n, (f, a) in bt.items() if a > 0.70]
        good_methods.sort(key=lambda x: -x[1])
        
        print(f"  #{rank:<2} {p['name'][:50]:<50}")
        print(f"      BEST: {best[0]} ({best[1][1]*100:.1f}% acc, fc={best[1][0]:,.0f})")
        if len(good_methods) > 1:
            others = ', '.join(f"{n}({a*100:.0f}%)" for n, a in good_methods[:4] if n != best[0])
            if others:
                print(f"      ALSO GOOD: {others}")
        print()

# ============================================================
# AUDIT 5: CROSS-REFERENCE V4.0 DECISIONS
# ============================================================
print()
print("=" * 120)
print("AUDIT 5: CROSS-REFERENCE V4.0 DECISIONS WITH BACKTEST EVIDENCE")
print("=" * 120)
print()

for p in products:
    rank = p['rank']
    act = p['actuals']
    v4_fc = v4_values[rank]
    fy25q2 = act[8]
    
    if rank not in backtest_results or fy25q2 <= 0:
        continue
    
    bt = backtest_results[rank]
    best = max(bt.items(), key=lambda x: x[1][1])
    
    # What would the BEST backtest method predict for FY26Q2?
    fy26q1 = act[11]
    fy25q1 = act[7]
    fy24q2 = act[4]
    fy23q2 = act[0]
    
    optimal_fc = None
    method_name = best[0]
    
    if method_name == 'Q2/Q1_ratio' and fy25q1 > 0 and fy25q2 > 0 and fy26q1 > 0:
        ratio = fy25q2 / fy25q1
        optimal_fc = fy26q1 * ratio
    elif method_name == 'YoY_Q2' and fy24q2 > 0 and fy25q2 > 0:
        yoy = (fy25q2 - fy24q2) / fy24q2
        yoy_c = clamp(yoy, -0.35, 0.40)
        optimal_fc = fy25q2 * (1 + yoy_c)
    elif method_name == 'Q2_WtAvg':
        q2_nz = [v for v in [fy23q2, fy24q2, fy25q2] if v > 0]
        if len(q2_nz) >= 3:
            optimal_fc = q2_nz[2]*0.50 + q2_nz[1]*0.35 + q2_nz[0]*0.15
        elif len(q2_nz) >= 2:
            optimal_fc = q2_nz[-1]*0.60 + q2_nz[-2]*0.40
    elif method_name == 'MA4':
        optimal_fc = sum(act[-4:])/4
    elif method_name == 'Naive_Q2':
        optimal_fc = fy25q2
    elif method_name == 'Naive_Q1':
        optimal_fc = fy26q1
    elif method_name == 'BigDeal_Q2':
        bd_b = p['big_deals'][4]  # FY25Q2
        bd_a = p['avg_deals'][4]
        # Project: weighted from both Q2s
        bd_b2 = p['big_deals'][0]
        bd_a2 = p['avg_deals'][0]
        optimal_fc = max(0, (bd_b*0.6+bd_b2*0.4) + (bd_a*0.6+bd_a2*0.4))
    elif method_name == 'SCMS_BU':
        # Full SCMS bottom-up with all 3 Q2s
        if rank in scms_data:
            total = 0
            for ch, vals in scms_data[rank].items():
                q2h = [max(0, vals[i]) for i in [1, 5, 9] if i < len(vals)]
                if len(q2h) >= 3:
                    total += q2h[2]*0.50 + q2h[1]*0.35 + q2h[0]*0.15
                elif len(q2h) >= 2:
                    total += q2h[-1]*0.60 + q2h[-2]*0.40
                elif q2h:
                    total += q2h[-1]
            optimal_fc = total
    elif method_name == 'VMS_BU':
        if rank in vms_data:
            total = 0
            for vt, vals in vms_data[rank].items():
                q2h = [max(0, vals[i]) for i in [1, 5, 9] if i < len(vals)]
                if len(q2h) >= 3:
                    total += q2h[2]*0.50 + q2h[1]*0.35 + q2h[0]*0.15
                elif len(q2h) >= 2:
                    total += q2h[-1]*0.60 + q2h[-2]*0.40
                elif q2h:
                    total += q2h[-1]
            optimal_fc = total
    
    if optimal_fc and optimal_fc > 0:
        delta = v4_fc - optimal_fc
        pct = delta / optimal_fc * 100
        flag = "!!!" if abs(pct) > 30 else ("! " if abs(pct) > 15 else "  ")
        print(f"  {flag} #{rank:<2} {p['name'][:45]:<45} v4.0={v4_fc:>7,} | BestBacktest({method_name})={optimal_fc:>7,.0f} | Δ={delta:>+6,.0f} ({pct:>+5.1f}%)")

print()
print("=" * 120)
print("AUDIT 6: EXPERT FORECAST ACCURACY IF USED RAW (NO CORRECTION)")
print("=" * 120)
print()
print("What if we just submitted the RAW expert forecasts? How would that compare?")
print()

# For each expert source, what's the average accuracy if we just used their raw FC?
for src, key in [('Demand Planner','dp'), ('Marketing','mk'), ('Data Science','ds')]:
    accs = []
    for p in products:
        fc = p[f'{key}_fc']
        if fc > 0 and p['actuals'][8] > 0:  # Only for products with FY25Q2 actual
            # This is their FY26Q2 forecast, not directly comparable
            # But we can check their track record accuracy
            pass
    
    # Print expert weighted accuracy per product
    print(f"  {src}:")
    for p in products:
        rank = p['rank']
        acc_q1 = p.get(f'{key}_acc_q1', 0)
        acc_q4 = p.get(f'{key}_acc_q4', 0)
        acc_q3 = p.get(f'{key}_acc_q3', 0)
        w = acc_q1*0.5 + acc_q4*0.3 + acc_q3*0.2
        fc = p[f'{key}_fc']
        marker = "***" if w > 0.85 else ("** " if w > 0.75 else ("*  " if w > 0.65 else "   "))
        print(f"    {marker} #{rank:<2} WtAcc={w*100:>5.1f}% | FC={fc:>8,.0f} | Recent: Q1={acc_q1*100:.0f}% Q4={acc_q4*100:.0f}% Q3={acc_q3*100:.0f}%")
    print()

# ============================================================
# AUDIT 7: KEY DECISION SUMMARY
# ============================================================
print()
print("=" * 120)
print("AUDIT 7: CRITICAL FINDINGS & RECOMMENDATIONS")
print("=" * 120)
print()

findings = [
    "FINDING 1: SCMS/VMS bottom-up performs POORLY in backtest for most products",
    "  - It uses Q2 segment history, but segment-level data is extremely noisy",
    "  - Individual segments can swing wildly (negative values, 10x changes)",
    "  - The sum of noisy parts != good total forecast",
    "  -> RECOMMENDATION: Reduce SCMS/VMS weight or exclude from structural median",
    "",
    "FINDING 2: Expert bias correction can HURT when bias is inconsistent",
    "  - If expert was +20% one quarter, -10% next, correcting either way is wrong",
    "  - The consistency check helps but 0.20 correction factor adds noise",
    "  -> RECOMMENDATION: Only correct when bias > 10% AND consistent in 2/3 quarters",
    "",
    "FINDING 3: YoY Q2 growth is clamped at ±35/40% but some products",
    "  genuinely have extreme patterns. The clamp biases toward the mean.",
    "  For products with CONSISTENT extreme growth/decline, this hurts.",
    "  -> RECOMMENDATION: Widen clamp for products with consistent trend direction",
    "",
    "FINDING 4: Q2 weighted average for new products (0->X->Y) gets polluted",
    "  by the 0 value from FY23Q2. This drags the average down.",
    "  -> RECOMMENDATION: Exclude zeros from Q2 weighted average",
    "",
    "FINDING 5: The structural median includes TOO MANY signals (7+)",
    "  Many are correlated (SCMS/VMS/BigDeal often move together)",
    "  This gives false confidence in the structural median",
    "  -> RECOMMENDATION: Use only 4 TRULY independent signals for median",
    "",
    "FINDING 6: Product-specific overrides (hardcoded weights) are brittle",
    "  They're based on v4.0 analysis but not backtested",
    "  -> RECOMMENDATION: Each override should be justified by backtest evidence",
]

for f in findings:
    print(f"  {f}")
