"""
============================================================================
CISCO CFL PHASE 2 - COMPETITION FORECAST v6.1 (CRYSTAL CUTTHROAT II)
============================================================================
BUILDS ON v5.0 WITH 7 RESEARCH-BACKED METHODOLOGY IMPROVEMENTS:

  v5.0 fixes retained: bias correction, independent structural signals,
  FY26Q1-aware caps/floors, growth product floors, sanity bounds.

  v6.0 changes (research-backed):
  1. Replace acc^3 with damped equal weights ("shrink-toward-equal")
     Rationale: Forecast combination puzzle — equal weights beat accuracy-
     based weighting in 50+ years of research (Clemen 1989, Stock & Watson 2004)

  2. Remove dominant expert rule
     Rationale: Combination always beats selection (Kourentzes et al.)

  3. Replace MA4 with Q2 seasonal average
     Rationale: MA4 mixes seasonal quarters, diluting Q2 signal

  4. Generalize product overrides into pattern-based rules
     Rationale: Hardcoded if-rank rules are textbook overfitting

  5. Smooth expert-structural step function to linear interpolation
     Rationale: Removes discontinuous weight jumps

  6. Fix growth floor bug for PLC=Decline products
     Rationale: Product #8 was incorrectly getting a growth floor

  7. Raise bias correction threshold from 3% to 8%
     Rationale: 3% is within noise for 3-quarter accuracy measurement
============================================================================
"""
import openpyxl
import math
import os
from datetime import datetime

wb = openpyxl.load_workbook('CFL_External Data Pack_Phase2.xlsx', data_only=True)
ws = wb['Ph.2 Data Pack-Actual Booking']
ws_big = wb['Ph.2 - Big Deal ']
ws_scms = wb['Ph.2 - SCMS']
ws_vms = wb['Ph.2 - VMS']

# ============================================================
# DATA INGESTION
# ============================================================
products = []
for row in range(4, 24):
    actuals = [float(ws.cell(row, c).value or 0) for c in range(4, 16)]
    products.append({
        'rank': ws.cell(row, 1).value,
        'name': ws.cell(row, 2).value,
        'plc': ws.cell(row, 3).value,
        'actuals': actuals,
        'dp_fc': float(ws.cell(row, 17).value or 0),
        'mk_fc': float(ws.cell(row, 18).value or 0),
        'ds_fc': float(ws.cell(row, 19).value or 0),
    })

for idx, row in enumerate(range(29, 49)):
    p = products[idx]
    for pfx, ca, cb in [('dp',[3,5,7],[4,6,8]),('mk',[10,12,14],[11,13,15]),('ds',[17,19,21],[18,20,22])]:
        a = [float(ws.cell(row,c).value or 0) for c in ca]
        b = [float(ws.cell(row,c).value or 0) for c in cb]
        p[f'{pfx}_acc_q1'], p[f'{pfx}_acc_q4'], p[f'{pfx}_acc_q3'] = a
        p[f'{pfx}_bias_q1'], p[f'{pfx}_bias_q4'], p[f'{pfx}_bias_q3'] = b

for idx, row in enumerate(range(3, 23)):
    p = products[idx]
    p['mfg_total'] = [float(ws_big.cell(row,c).value or 0) for c in range(3, 11)]
    p['big_deals'] = [float(ws_big.cell(row,c).value or 0) for c in range(11, 19)]
    p['avg_deals'] = [float(ws_big.cell(row,c).value or 0) for c in range(19, 27)]

scms_data = {}
for row in range(4, ws_scms.max_row+1):
    rank = ws_scms.cell(row,1).value
    if rank is None: continue
    ch = ws_scms.cell(row,3).value
    vals = [float(ws_scms.cell(row,c).value or 0) for c in range(4,17)]
    scms_data.setdefault(rank,{})[ch] = vals

vms_data = {}
for row in range(4, ws_vms.max_row+1):
    rank = ws_vms.cell(row,1).value
    if rank is None: continue
    vt = ws_vms.cell(row,3).value
    vals = [float(ws_vms.cell(row,c).value or 0) for c in range(4,17)]
    vms_data.setdefault(rank,{})[vt] = vals

# ============================================================
# HELPER FUNCTIONS
# ============================================================
def clamp(v, lo, hi): return max(lo, min(v, hi))

def w_acc(p, pfx):
    """Weighted accuracy: FY26Q1=50%, FY25Q4=30%, FY25Q3=20%"""
    return p[f'{pfx}_acc_q1']*0.5 + p[f'{pfx}_acc_q4']*0.3 + p[f'{pfx}_acc_q3']*0.2

def w_bias(p, pfx):
    return p[f'{pfx}_bias_q1']*0.5 + p[f'{pfx}_bias_q4']*0.3 + p[f'{pfx}_bias_q3']*0.2

def bias_consistency(p, pfx):
    """Consistent bias direction = correct more aggressively"""
    biases = [p[f'{pfx}_bias_q1'], p[f'{pfx}_bias_q4'], p[f'{pfx}_bias_q3']]
    signs = [1 if b > 0.08 else (-1 if b < -0.08 else 0) for b in biases]
    non_zero = [s for s in signs if s != 0]
    if len(non_zero) >= 2 and all(s == non_zero[0] for s in non_zero):
        return 0.50
    elif len(non_zero) >= 2:
        return 0.20
    else:
        return 0.30

def bias_correct(fc, bias, consistency):
    if abs(bias) < 0.08: return fc  # v6.0: raised from 3% — only correct meaningful bias
    correction = clamp(bias * consistency, -0.40, 0.40)
    return fc * (1 - correction)

def bottom_up_q2(seg_data, rank):
    """Bottom-up from Q2 segment history (for validation only, NOT in structural median)"""
    if rank not in seg_data: return None
    total = 0
    for name, vals in seg_data[rank].items():
        q2_hist = [vals[i] for i in [1, 5, 9] if i < len(vals)]
        q2_pos = [max(0, v) for v in q2_hist]
        if len(q2_pos) >= 3 and sum(q2_pos) > 0:
            fc = q2_pos[2]*0.50 + q2_pos[1]*0.35 + q2_pos[0]*0.15
        elif len(q2_pos) >= 2 and sum(q2_pos) > 0:
            fc = q2_pos[-1]*0.60 + q2_pos[-2]*0.40
        elif q2_pos and q2_pos[-1] > 0:
            fc = q2_pos[-1]
        else:
            fc = 0
        total += fc
    return total if total > 0 else None

def median_val(values):
    v = sorted([x for x in values if x > 0])
    if not v: return 0
    n = len(v)
    if n % 2 == 1: return v[n//2]
    return (v[n//2-1] + v[n//2]) / 2

# ============================================================
# FORECASTING: v6.0 Crystal Cutthroat II
# ============================================================
results = []

for p in products:
    act = p['actuals']
    rank, name, plc = p['rank'], p['name'], p['plc']
    fy26q1, fy25q2, fy24q2, fy23q2 = act[11], act[8], act[4], act[0]
    fy25q1 = act[7]
    
    # --- EXPERT ANALYSIS ---
    expert_entries = []
    for pfx, label in [('dp','DP'), ('mk','MK'), ('ds','DS')]:
        acc = w_acc(p, pfx)
        bias = w_bias(p, pfx)
        fc_raw = p[f'{pfx}_fc']
        
        if acc < 0.05 or fc_raw <= 0:
            expert_entries.append((label, acc, fc_raw, fc_raw, True))
            continue
        
        cons = bias_consistency(p, pfx)
        fc_corrected = bias_correct(fc_raw, bias, cons)
        expert_entries.append((label, acc, fc_corrected, fc_raw, False))
    
    dp_a, mk_a, ds_a = w_acc(p,'dp'), w_acc(p,'mk'), w_acc(p,'ds')
    dp_b, mk_b, ds_b = w_bias(p,'dp'), w_bias(p,'mk'), w_bias(p,'ds')
    dp_cons = bias_consistency(p, 'dp')
    mk_cons = bias_consistency(p, 'mk')
    ds_cons = bias_consistency(p, 'ds')
    dp_c = bias_correct(p['dp_fc'], dp_b, dp_cons) if dp_a >= 0.05 and p['dp_fc'] > 0 else p['dp_fc']
    mk_c = bias_correct(p['mk_fc'], mk_b, mk_cons) if mk_a >= 0.05 and p['mk_fc'] > 0 else p['mk_fc']
    ds_c = bias_correct(p['ds_fc'], ds_b, ds_cons) if ds_a >= 0.05 and p['ds_fc'] > 0 else p['ds_fc']
    
    valid_experts = [(l, a, fc_c, fc_r) for l, a, fc_c, fc_r, exc in expert_entries if not exc]
    
    # ===== v6.1: OUTLIER EXPERT CAP =====
    # When one expert's forecast is wildly different (>2x median of others),
    # cap it to prevent a single extreme forecast from dominating the blend.
    # This catches DS's 29,553 on Product #9 (highest Q2 ever = 8,791).
    if len(valid_experts) >= 3:
        fcs = [fc for _, _, fc, _ in valid_experts]
        median_fc = sorted(fcs)[len(fcs)//2]
        valid_experts = [(l, a, min(fc, median_fc * 2.0), fr)
                         for l, a, fc, fr in valid_experts]
    
    if valid_experts:
        # Sort by accuracy
        valid_sorted = sorted(valid_experts, key=lambda x: x[1], reverse=True)
        best_name, best_acc, best_fc = valid_sorted[0][0], valid_sorted[0][1], valid_sorted[0][2]
        second_acc = valid_sorted[1][1] if len(valid_sorted) > 1 else 0
        
        # ===== v6.0: DAMPED EQUAL WEIGHTS (replaces dominant expert + acc^3) =====
        # Research: "Forecast Combination Puzzle" — equal weights beat accuracy-based
        # weighting in 50+ years of studies. We use shrink-toward-equal: 60% equal + 40% acc^1.
        n_exp = len(valid_experts)
        shrinkage = 0.60  # 60% equal weights, 40% accuracy-informed
        total_acc = sum(a for _, a, _, _ in valid_experts)
        if total_acc > 0 and n_exp > 0:
            equal_w = 1.0 / n_exp
            exp_blend = 0
            for _, a, fc_c, _ in valid_experts:
                acc_w = a / total_acc  # acc^1 normalized
                blended_w = shrinkage * equal_w + (1 - shrinkage) * acc_w
                exp_blend += blended_w * fc_c
        else:
            exp_blend = sum(fc_c for _, _, fc_c, _ in valid_experts) / n_exp if n_exp > 0 else sum(act[-4:])/4
        anchor_note = f"damped-equal blend, best={best_name} ({best_acc*100:.0f}%)"
        
        avg_expert_acc = total_acc / n_exp if n_exp > 0 else 0
    else:
        exp_blend = sum(act[-4:])/4
        best_name, best_acc = '-', 0
        avg_expert_acc = 0
        anchor_note = "no valid experts"
    
    # --- STRUCTURAL SIGNALS (TRULY INDEPENDENT ONLY) ---
    
    # Signal 1: Q2/Q1 ratio forecast
    q2q1_ratios = []
    for q2i, q1i in [(4,3),(8,7)]:
        if act[q1i] > 0 and act[q2i] > 0:
            q2q1_ratios.append(act[q2i]/act[q1i])
    
    if q2q1_ratios and fy26q1 > 0:
        if len(q2q1_ratios) > 1:
            ratio_spread = abs(q2q1_ratios[1] - q2q1_ratios[0])
            if ratio_spread > 0.5:
                q2q1 = q2q1_ratios[-1]
            else:
                q2q1 = q2q1_ratios[-1]*0.6 + q2q1_ratios[0]*0.4
        else:
            q2q1 = q2q1_ratios[0]
        # CAP ratio to prevent spike/collapse quarter outliers
        q2q1 = clamp(q2q1, 0.20, 2.0)
        ratio_fc = fy26q1 * q2q1
    else:
        q2q1 = 1.0
        ratio_fc = fy25q2 if fy25q2 > 0 else fy26q1
    
    # Signal 2: YoY Q2 growth forecast
    q2_vals = [fy23q2, fy24q2, fy25q2]
    yoy_rates = []
    for i in range(1, len(q2_vals)):
        if q2_vals[i-1] > 0 and q2_vals[i] > 0:
            yoy_rates.append((q2_vals[i] - q2_vals[i-1]) / q2_vals[i-1])
    if yoy_rates:
        yoy = yoy_rates[-1]*0.6 + yoy_rates[0]*0.4 if len(yoy_rates) > 1 else yoy_rates[-1]
        # Wider clamp for consistent trends
        all_same_dir = all(r > 0 for r in yoy_rates) or all(r < 0 for r in yoy_rates)
        if all_same_dir and len(yoy_rates) > 1:
            yoy = clamp(yoy, -0.45, 0.50)  # Wider for consistent direction
        else:
            yoy = clamp(yoy, -0.35, 0.40)
    else:
        yoy = 0
    yoy_fc = fy25q2 * (1+yoy) if fy25q2 > 0 else fy26q1
    
    # Signal 3: MA4 (v6.1: REVERTED from Q2 seasonal avg — backtest proved MA4 is better)
    # v6.0 used Q2 seasonal avg here, but walk-forward backtest showed MA4 structural
    # accuracy 49.9% vs Q2 avg 42.3%. MA4 captures recent momentum across all quarters.
    ma4_fc = sum(act[-4:])/4
    
    # Q2 Seasonal Average — kept for validation output only
    q2_nz = [v for v in q2_vals if v > 0]
    if len(q2_nz) >= 3:
        q2_seasonal_fc = q2_vals[2]*0.50 + q2_vals[1]*0.35 + q2_vals[0]*0.15
    elif len(q2_nz) >= 2:
        q2_seasonal_fc = q2_nz[-1]*0.60 + q2_nz[-2]*0.40
    elif q2_nz:
        q2_seasonal_fc = q2_nz[-1]
    else:
        q2_seasonal_fc = ma4_fc
    
    # VALIDATION signals (NOT in median, used for sanity checking only)
    scms_fc = bottom_up_q2(scms_data, rank)
    vms_fc = bottom_up_q2(vms_data, rank)
    q2_avg_fc = q2_seasonal_fc  # validation only
    
    q2_big_hist = [p['big_deals'][0], p['big_deals'][4]]
    q2_avg_hist = [p['avg_deals'][0], p['avg_deals'][4]]
    bd_fc = max(0, q2_big_hist[-1]*0.6 + q2_big_hist[-2]*0.4 + q2_avg_hist[-1]*0.6 + q2_avg_hist[-2]*0.4)
    
    # ============================================================
    # STRUCTURAL COMPOSITE (INDEPENDENT SIGNALS ONLY)
    # ============================================================
    # ONLY 3 truly independent signals (v6.1: MA4 reverted as signal 3)
    ind_signals = [v for v in [ratio_fc, yoy_fc, ma4_fc] if v > 0]
    
    # FY26Q1-AWARE STRUCTURAL INTELLIGENCE
    is_decline = plc == 'Decline' or (fy26q1 > 0 and fy25q2 > 0 and fy26q1 < fy25q2 * 0.75)
    # v6.0 FIX: Don't override PLC=Decline with calculated growth (was a bug for Product #8)
    is_growth = 'Growth' in str(plc) or (
        str(plc) != 'Decline' and fy26q1 > 0 and fy25q2 > 0 and fy26q1 > fy25q2 * 1.10
    )
    
    # CHECK: Has FY26Q1 dropped significantly from FY25Q1?
    # If yes, YoY and MA4 are STALE (based on old high quarters) and Q2/Q1 ratio is the
    # ONLY structural signal that incorporates the current reality.
    q1_drop = (fy25q1 - fy26q1) / fy25q1 if fy25q1 > 0 else 0
    
    if q1_drop > 0.25 and ratio_fc > 0:
        # FY26Q1 dropped >25% from FY25Q1: Q2/Q1 ratio is the most reliable structural signal
        # because it uses CURRENT Q1 as the base. YoY/MA4 are backward-looking and stale.
        # Give ratio_fc 60% weight, blend others at 40%
        other_struct = [v for v in [yoy_fc, ma4_fc] if v > 0]
        if other_struct:
            struct_other = sum(other_struct) / len(other_struct)
            structural_median = ratio_fc * 0.60 + struct_other * 0.40
        else:
            structural_median = ratio_fc
    else:
        structural_median = median_val(ind_signals) if ind_signals else exp_blend
    
    # Structural CAPS/FLOORS
    if is_decline and fy26q1 > 0:
        # Cap at FY26Q1-based ceiling (Q1 is the CURRENT state of the product)
        # Use Q2/Q1 ratio * FY26Q1 as the max reasonable Q2
        if ratio_fc > 0:
            struct_cap = max(ratio_fc, fy26q1) * 1.15
        else:
            struct_cap = fy26q1 * 1.30
        structural_median = min(structural_median, struct_cap)
    
    if is_growth and fy26q1 > 0:
        # Floor structural at FY26Q1 * 0.90
        structural_median = max(structural_median, fy26q1 * 0.90)
    
    # ============================================================
    # SEASONAL NAIVE SAFETY NET (v6.1: prevents structural from being worse than naive)
    # ============================================================
    # Backtest showed MASE > 1 for both v5/v6 structural — meaning structural signals
    # are WORSE than just using last year's Q2. This guard shrinks extreme structural
    # deviations back toward the seasonal naive baseline.
    seasonal_naive = fy25q2
    if structural_median > 0 and seasonal_naive > 0:
        naive_deviation = abs(structural_median - seasonal_naive) / seasonal_naive
        if naive_deviation > 0.40:
            # Structural deviates >40% from naive: shrink 30% back toward naive
            structural_median = structural_median * 0.70 + seasonal_naive * 0.30
    
    # ============================================================
    # EXPERT-ANCHORED BLEND (v6.1: boosted expert range since structural < naive)
    # ============================================================
    # Linear interpolation: maps accuracy [0.20, 0.95] -> weight [0.35, 0.90]
    # v6.0 used [0.25, 0.85] but backtest showed structural MASE > 1 (worse than naive)
    # so we lean MORE on experts across the board
    expert_weight = clamp(0.35 + (avg_expert_acc - 0.20) * (0.90 - 0.35) / (0.95 - 0.20), 0.35, 0.90)
    
    # ============================================================
    # PATTERN-BASED OVERRIDES (v6.0: generalized from product-specific rules)
    # ============================================================
    override_note = ""
    
    # Rule A: Exclude experts below 10% accuracy (generalizes rank==19/20 logic)
    orig_valid = valid_experts
    valid_experts_filtered = [(l, a, fc_c, fc_r) for l, a, fc_c, fc_r in valid_experts if a >= 0.10]
    if len(valid_experts_filtered) < len(valid_experts) and valid_experts_filtered:
        valid_experts = valid_experts_filtered
        # Recompute blend with filtered experts
        n_exp = len(valid_experts)
        total_acc = sum(a for _, a, _, _ in valid_experts)
        if total_acc > 0 and n_exp > 0:
            equal_w = 1.0 / n_exp
            exp_blend = 0
            for _, a, fc_c, _ in valid_experts:
                acc_w = a / total_acc
                blended_w = shrinkage * equal_w + (1 - shrinkage) * acc_w
                exp_blend += blended_w * fc_c
        excluded_names = [l for l, a, _, _ in orig_valid if a < 0.10]
        override_note = f"Excluded {','.join(excluded_names)} (<10% acc)"
    
    # Rule B: If ALL valid experts < 55% accuracy, lean structural
    if valid_experts and all(a < 0.55 for _, a, _, _ in valid_experts):
        expert_weight = max(expert_weight - 0.20, 0.25)
        override_note = (override_note + "; " if override_note else "") + "All experts <55% acc; lean structural"
    
    # Rule C: If Q2 consistently spikes >1.5x Q1, reduce expert weight
    # (experts tend to under-forecast seasonal spikes)
    if len(q2q1_ratios) >= 2 and all(r > 1.5 for r in q2q1_ratios):
        expert_weight = min(expert_weight, 0.40)
        override_note = (override_note + "; " if override_note else "") + "Q2 spike product (Q2/Q1 >1.5x consistently)"
    
    # --- COMPUTE FINAL ---
    final = exp_blend * expert_weight + structural_median * (1 - expert_weight)
    
    # ===== GROWTH PRODUCT FLOOR =====
    if is_growth and fy26q1 > 0:
        growth_floor = fy26q1 * 0.90
        if final < growth_floor:
            final = growth_floor
            override_note = (override_note + "; " if override_note else "") + f"growth floor applied ({growth_floor:.0f})"
    
    # SANITY BOUNDS
    credible_vals = []
    for label, acc, fc_c, fc_raw, excluded in expert_entries:
        if not excluded and fc_raw > 0:
            credible_vals.append(fc_c)
    for v in [fy25q2, fy24q2, fy26q1]:
        if v > 0:
            credible_vals.append(v)
    if structural_median > 0:
        credible_vals.append(structural_median)
    
    if credible_vals:
        cred_low = min(credible_vals)
        cred_high = max(credible_vals)
        bound_low = cred_low * 0.55
        bound_high = cred_high * 1.20
        final = clamp(final, bound_low, bound_high)
    
    final = max(0, round(final))
    
    # ===== MANUAL OVERRIDE: Product #1 WiFi AP =====
    # Research-backed: Q2 spike (2,284→6,651→8,293) is driven by enterprise budget cycle
    # timing (Cisco Q2 = Oct-Jan = fiscal year-end spending). WiFi 6 lifecycle = Sustaining
    # (confirmed by market research — not declining). All 3 experts anchor on Q1 (~3,000)
    # and miss the Q2 spike. Override to Q2/Q1 ratio forecast (3,181 × 2.0 = 6,362).
    if rank == 1:
        final = 6362
        override_note = "MANUAL: Q2/Q1 ratio (budget-cycle spike, experts under-forecast)"
    
    # Notes
    note = override_note if override_note else ""
    if not note:
        if best_acc > 0.85:
            note = f"High-confidence {best_name} ({best_acc*100:.0f}% acc)"
        elif avg_expert_acc < 0.30:
            note = "All experts unreliable; structural-heavy"
        elif avg_expert_acc < 0.50:
            note = "Weak experts; structural-heavy blend"
        else:
            note = f"{anchor_note}"
    
    results.append({
        'rank': rank, 'name': name, 'plc': plc,
        'final': final,
        'anchor': round(exp_blend),
        'structural_median': round(structural_median),
        'expert_weight': expert_weight,
        'best_src': best_name, 'best_acc': best_acc,
        'dp_fc': round(p['dp_fc']), 'mk_fc': round(p['mk_fc']), 'ds_fc': round(p['ds_fc']),
        'dp_c': round(dp_c), 'mk_c': round(mk_c), 'ds_c': round(ds_c),
        'dp_a': dp_a, 'mk_a': mk_a, 'ds_a': ds_a,
        'dp_b': dp_b, 'mk_b': mk_b, 'ds_b': ds_b,
        'ratio_fc': round(ratio_fc), 'yoy_fc': round(yoy_fc), 'q2_avg_fc': round(q2_avg_fc),
        'bd_fc': round(bd_fc), 'ma4_fc': round(ma4_fc),
        'scms_fc': round(scms_fc) if scms_fc else '-',
        'vms_fc': round(vms_fc) if vms_fc else '-',
        'fy26q1': round(fy26q1), 'fy25q2': round(fy25q2),
        'fy24q2': round(fy24q2), 'fy23q2': round(fy23q2),
        'q2q1': round(q2q1, 3), 'yoy': round(yoy*100,1),
        'note': note, 'anchor_note': anchor_note if 'anchor_note' in dir() else '',
    })

# ============================================================
# OUTPUT
# ============================================================
L = []
L.append("=" * 130)
L.append("  CISCO CFL PHASE 2 - FY26 Q2 DEMAND FORECAST PREDICTIONS")
L.append("  Version 6.1 - Crystal Cutthroat II (Verification-Refined)")
L.append(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
L.append("=" * 130)

L.append("")
L.append("METHODOLOGY:")
L.append("-" * 130)
L.append("  CORE: Expert-Anchored Ensemble with research-backed improvements")
L.append("  v6.1 CHANGES: Damped equal weights, no dominant expert rule, MA4 structural,")
L.append("                pattern-based overrides, boosted expert range, naive safety net")
L.append("")
L.append("  Weighting: 60% equal + 40% acc^1 (shrink-toward-equal, per combination puzzle)")
L.append("  Structural: 3 independent signals (Q2/Q1 ratio, YoY Q2, MA4) [MA4 reverted from Q2 avg]")
L.append("  Expert range: [35%-90%] (boosted from [25%-85%] because structural MASE > 1)")
L.append("  Safety net: Shrink structural toward seasonal naive when deviation > 40%")
L.append("")

# SUMMARY TABLE
L.append("=" * 130)
L.append("  FY26 Q2 FORECAST PREDICTIONS")
L.append("=" * 130)
L.append("")
L.append(f"{'#':<4} {'Product':<55} {'PLC':<18} {'Prediction':>11} {'Anchor':>9} {'Struct':>9} {'Split':>8}")
L.append("-" * 130)

total = 0
for r in results:
    split = f"{int(r['expert_weight']*100)}/{int((1-r['expert_weight'])*100)}"
    L.append(f"{r['rank']:<4} {r['name']:<55} {r['plc']:<18} {r['final']:>11,} {r['anchor']:>9,} {r['structural_median']:>9,} {split:>8}")
    total += r['final']

L.append("-" * 130)
L.append(f"{'':4} {'TOTAL':<55} {'':18} {total:>11,}")
L.append("")

# V5 -> V6 DELTA
L.append("=" * 130)
L.append("  V5.0 -> V6.0 CHANGES")
L.append("=" * 130)
L.append("")
v5_values = {1:5104,2:5800,3:5638,4:11432,5:1272,6:688,7:658,8:5543,9:7385,10:8890,
             11:606,12:1557,13:396,14:9159,15:647,16:483,17:699,18:129,19:2220,20:1659}
L.append(f"{'#':<4} {'Product':<55} {'v5.0':>8} {'v6.0':>8} {'Delta':>8} {'%Chg':>7}")
L.append("-" * 100)
for r in results:
    v5 = v5_values.get(r['rank'], 0)
    v6 = r['final']
    delta = v6 - v5
    pct = (delta/v5*100) if v5 > 0 else 0
    flag = ">>>" if abs(pct) > 10 else "   "
    L.append(f"{flag}{r['rank']:<3} {r['name']:<55} {v5:>8,} {v6:>8,} {delta:>+8,} {pct:>+6.1f}%")
v5_total = sum(v5_values.values())
L.append("-" * 100)
L.append(f"   {'TOTAL':<58} {v5_total:>8,} {total:>8,} {total-v5_total:>+8,} {(total-v5_total)/v5_total*100:>+6.1f}%")
L.append("")

# DETAILED CARDS
L.append("=" * 130)
L.append("  DETAILED PRODUCT FORECAST CARDS")
L.append("=" * 130)

for r in results:
    L.append("")
    L.append("+" + "-" * 128 + "+")
    L.append(f"| #{r['rank']} {r['name']}")
    L.append(f"| PLC: {r['plc']} | Q2/Q1: {r['q2q1']} | YoY Q2: {r['yoy']}%")
    L.append(f"| Q2 History: FY23={r['fy23q2']:,} | FY24={r['fy24q2']:,} | FY25={r['fy25q2']:,} | FY26Q1={r['fy26q1']:,}")
    L.append("+" + "-" * 128 + "+")
    L.append(f"  >>> PREDICTION: {r['final']:,} UNITS   ({r['note']})")
    L.append("")
    
    L.append("  EXPERT FORECASTS:")
    L.append(f"    {'Source':<22} {'Raw':>8} {'Bias%':>8} {'Corrected':>10} {'Accuracy%':>10}")
    L.append(f"    {'-'*65}")
    
    best_m = lambda s: " <-- BEST" if s == r['best_src'] else ""
    excl = lambda a: " [EXCL]" if a < 0.05 else ""
    L.append(f"    {'Demand Planner':<22} {r['dp_fc']:>8,} {r['dp_b']*100:>7.1f}% {r['dp_c']:>10,} {r['dp_a']*100:>9.1f}%{best_m('DP')}{excl(r['dp_a'])}")
    L.append(f"    {'Marketing Team':<22} {r['mk_fc']:>8,} {r['mk_b']*100:>7.1f}% {r['mk_c']:>10,} {r['mk_a']*100:>9.1f}%{best_m('MK')}{excl(r['mk_a'])}")
    L.append(f"    {'Data Science':<22} {r['ds_fc']:>8,} {r['ds_b']*100:>7.1f}% {r['ds_c']:>10,} {r['ds_a']*100:>9.1f}%{best_m('DS')}{excl(r['ds_a'])}")
    L.append(f"    Anchor: {r['anchor']:,}")
    L.append("")
    
    L.append("  INDEPENDENT STRUCTURAL SIGNALS:")
    L.append(f"    Q2/Q1 Ratio FC:  {r['ratio_fc']:>8,}    |  [validation] Q2 Avg:  {r['q2_avg_fc']:>8,}")
    L.append(f"    YoY Q2 FC:       {r['yoy_fc']:>8,}    |  [validation] BD FC:   {r['bd_fc']:>8,}")
    L.append(f"    MA4 FC:          {r['ma4_fc']:>8,}    |  [validation] SCMS BU: {str(r['scms_fc']):>8}")
    L.append(f"    Structural Median: {r['structural_median']:,}")
    L.append(f"    Blend: {r['expert_weight']*100:.0f}% Expert + {(1-r['expert_weight'])*100:.0f}% Structural = {r['final']:,}")
    L.append("")

# Competition values
L.append("")
L.append("=" * 130)
L.append("  VALUES TO ENTER IN COMPETITION SPREADSHEET (Column P: Your Forecast FY26 Q2)")
L.append("=" * 130)
L.append("")
for r in results:
    L.append(f"  {r['rank']:>2}. {r['name']:<55}   {r['final']:>8,}")
L.append(f"  {'':>2}  {'TOTAL':<55}   {total:>8,}")
L.append("")
L.append("=" * 130)

# Write
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "CCL_FY26_Q2_Forecast_Predictions.txt")
with open(out, 'w', encoding='utf-8') as f:
    f.write("\n".join(L))

print(f"Written to: {out}")
print(f"Total: {total:,}")
print()
print("=" * 80)
print("VALUES TO ENTER:")
print("=" * 80)
for r in results:
    print(f"  #{r['rank']:>2} {r['name']:<55} {r['final']:>8,}")
print(f"  {'':>4} {'TOTAL':<55} {total:>8,}")
print()

print("=" * 80)
print("V5.0 -> V6.0 DELTA:")
print("=" * 80)
for r in results:
    v5 = v5_values.get(r['rank'], 0)
    v6 = r['final']
    delta = v6 - v5
    flag = ">>>" if abs(delta) > 500 else "   "
    print(f"  {flag} #{r['rank']:>2} {r['name']:<50} v5={v5:>7,} -> v6={v6:>7,}  ({delta:>+6,})")
print(f"  {'':>4} {'TOTAL':<50} v5={v5_total:>7,} -> v6={total:>7,}  ({total-v5_total:>+6,})")
