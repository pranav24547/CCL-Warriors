import openpyxl

wb = openpyxl.load_workbook('Phase2AccuracyCalculation.xlsx', data_only=True)
print("Sheet names:", wb.sheetnames)
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== SHEET: {sheet_name} ===")
    print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")
    print()
    
    # Print all rows (up to 50)
    for row_idx in range(1, min(ws.max_row + 1, 50)):
        cells = []
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                cells.append(f"C{col_idx}={val}")
        if cells:
            print(f"  Row {row_idx}: {cells}")
    print()

wb.close()
